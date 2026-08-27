from __future__ import annotations

import csv
import hashlib
import io
import re
import os
import time
import uuid
import json
import base64
import urllib.request
import urllib.error
import hmac
import sqlite3
from collections import defaultdict
from datetime import datetime
from html import escape
from typing import Any

from fastapi import FastAPI, File, UploadFile, HTTPException, Request, Form
from fastapi.responses import HTMLResponse, JSONResponse, Response, RedirectResponse
from pydantic import BaseModel

app = FastAPI(title="СделатьВычет", version="2.4.2")


REPORT_PRICE_RUB = 499
ANALYSIS_TTL_SECONDS = 2 * 60 * 60
ANALYSES: dict[str, dict[str, Any]] = {}

YOOKASSA_SHOP_ID = os.getenv("YOOKASSA_SHOP_ID", "").strip()
YOOKASSA_SECRET_KEY = os.getenv("YOOKASSA_SECRET_KEY", "").strip()
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "").strip().rstrip("/")
PAYMENT_TEST_MODE = os.getenv("PAYMENT_TEST_MODE", "").strip().lower() in {"1", "true", "yes", "on"}


SERVICE_NAME = "СделатьВычет"
OPERATOR_NAME = "Колосов Роман Михайлович"
OPERATOR_INN = "772072450119"
OPERATOR_EMAIL = "inbox@sdelatvychet.ru"
# Закон требует адрес оператора в письменном согласии. До публичного запуска
# заполните эту переменную в Timeweb: OPERATOR_ADDRESS=...
OPERATOR_ADDRESS = os.getenv("OPERATOR_ADDRESS", "Зелёный проспект, д. 32").strip()

TERMS_VERSION = "2026-08-27-v1"
PRIVACY_VERSION = "2026-08-27-v1"
CONSENT_VERSION = "2026-08-27-v1"

# Аудит согласий. Для production путь должен находиться на постоянном диске/БД.
CONSENT_AUDIT_PATH = os.getenv("CONSENT_AUDIT_PATH", "data/consent_audit.jsonl").strip()
CONSENT_AUDIT_SECRET = os.getenv("CONSENT_AUDIT_SECRET", "").strip()


# Product analytics.
# No merchant names, transaction descriptions, exact transaction sums, account
# numbers or report contents are written to this database.
METRIKA_COUNTER_ID = os.getenv("METRIKA_COUNTER_ID", "").strip()
ANALYTICS_DB_PATH = os.getenv(
    "ANALYTICS_DB_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "analytics.sqlite3"),
).strip()
ANALYTICS_DASHBOARD_TOKEN = os.getenv("ANALYTICS_DASHBOARD_TOKEN", "").strip()

ANALYTICS_EVENTS = {
    "visit",
    "upload_click",
    "file_selected",
    "analysis_started",
    "analysis_success",
    "analysis_error",
    "result_view",
    "paywall_view",
    "payment_click",
    "payment_created",
    "payment_success",
    "payment_error",
    "report_view",
    "guide_step_1",
    "guide_step_2",
    "guide_step_3",
    "guide_complete",
}

def _analytics_connect():
    path = os.path.abspath(ANALYTICS_DB_PATH)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    con = sqlite3.connect(path, timeout=5)
    con.row_factory = sqlite3.Row
    return con

def init_analytics_db():
    try:
        with _analytics_connect() as con:
            con.execute("PRAGMA journal_mode=WAL")
            con.execute("""
                CREATE TABLE IF NOT EXISTS analytics_events (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL,
                    session_id TEXT NOT NULL,
                    event TEXT NOT NULL,
                    utm_source TEXT,
                    utm_medium TEXT,
                    utm_campaign TEXT,
                    utm_content TEXT,
                    utm_term TEXT,
                    referrer_host TEXT,
                    bank TEXT,
                    parser TEXT,
                    candidate_bucket TEXT,
                    refund_bucket TEXT,
                    payment_amount INTEGER,
                    dedupe_key TEXT UNIQUE
                )
            """)
            con.execute("CREATE INDEX IF NOT EXISTS idx_analytics_event_time ON analytics_events(event, created_at)")
            con.execute("CREATE INDEX IF NOT EXISTS idx_analytics_session ON analytics_events(session_id)")
            con.execute("CREATE INDEX IF NOT EXISTS idx_analytics_source ON analytics_events(utm_source)")
            con.execute("""
                CREATE TABLE IF NOT EXISTS public_stats (
                    id INTEGER PRIMARY KEY CHECK (id = 1),
                    total_found_refund REAL NOT NULL DEFAULT 0,
                    analyses_count INTEGER NOT NULL DEFAULT 0,
                    updated_at TEXT
                )
            """)
            con.execute("""
                INSERT OR IGNORE INTO public_stats (id, total_found_refund, analyses_count, updated_at)
                VALUES (1, 0, 0, ?)
            """, (datetime.utcnow().replace(microsecond=0).isoformat() + "Z",))
            con.commit()
    except Exception:
        # Analytics must never prevent the tax service itself from working.
        return False
    return True

def clean_analytics_value(value: Any, max_len: int = 120) -> str:
    s = str(value or "").strip()
    s = re.sub(r"[\x00-\x1f\x7f]", "", s)
    return s[:max_len]

def valid_session_id(value: str) -> bool:
    return bool(re.fullmatch(r"[a-zA-Z0-9_-]{16,80}", value or ""))

def candidate_bucket(n: int) -> str:
    if n <= 0: return "0"
    if n <= 2: return "1-2"
    if n <= 5: return "3-5"
    if n <= 10: return "6-10"
    if n <= 20: return "11-20"
    return "21+"

def refund_bucket(value: float) -> str:
    v = float(value or 0)
    if v <= 0: return "0"
    if v < 3000: return "<3k"
    if v < 10000: return "3-10k"
    if v < 20000: return "10-20k"
    if v < 50000: return "20-50k"
    return "50k+"

def record_analytics_event(
    *,
    event: str,
    session_id: str,
    attribution: dict[str, Any] | None = None,
    bank: str = "",
    parser: str = "",
    candidate_count: int | None = None,
    refund_value: float | None = None,
    payment_amount: int | None = None,
    dedupe_key: str | None = None,
) -> bool:
    if event not in ANALYTICS_EVENTS or not valid_session_id(session_id):
        return False
    a = attribution or {}
    row = (
        datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        session_id,
        event,
        clean_analytics_value(a.get("utm_source"), 80),
        clean_analytics_value(a.get("utm_medium"), 80),
        clean_analytics_value(a.get("utm_campaign"), 120),
        clean_analytics_value(a.get("utm_content"), 120),
        clean_analytics_value(a.get("utm_term"), 120),
        clean_analytics_value(a.get("referrer_host"), 120),
        clean_analytics_value(bank, 60),
        clean_analytics_value(parser, 120),
        candidate_bucket(candidate_count) if candidate_count is not None else "",
        refund_bucket(refund_value) if refund_value is not None else "",
        int(payment_amount) if payment_amount is not None else None,
        clean_analytics_value(dedupe_key, 180) if dedupe_key else None,
    )
    try:
        with _analytics_connect() as con:
            con.execute("""
                INSERT OR IGNORE INTO analytics_events (
                    created_at, session_id, event,
                    utm_source, utm_medium, utm_campaign, utm_content, utm_term,
                    referrer_host, bank, parser, candidate_bucket, refund_bucket,
                    payment_amount, dedupe_key
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, row)
            con.commit()
        return True
    except Exception:
        return False

def analytics_attribution_from_form(
    utm_source: str = "", utm_medium: str = "", utm_campaign: str = "",
    utm_content: str = "", utm_term: str = "", referrer_host: str = "",
) -> dict[str, str]:
    return {
        "utm_source": clean_analytics_value(utm_source, 80),
        "utm_medium": clean_analytics_value(utm_medium, 80),
        "utm_campaign": clean_analytics_value(utm_campaign, 120),
        "utm_content": clean_analytics_value(utm_content, 120),
        "utm_term": clean_analytics_value(utm_term, 120),
        "referrer_host": clean_analytics_value(referrer_host, 120),
    }

init_analytics_db()


def operator_address_display() -> str:
    return OPERATOR_ADDRESS or "АДРЕС ОПЕРАТОРА НЕ ЗАПОЛНЕН — задайте OPERATOR_ADDRESS перед публичным запуском"


def legal_ready() -> bool:
    return bool(OPERATOR_ADDRESS)


def request_ip(request: Request) -> str:
    forwarded = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    if forwarded:
        return forwarded
    return request.client.host if request.client else ""


def privacy_hash(value: str) -> str:
    """Store a pseudonymous proof rather than raw IP / user-agent."""
    key = (CONSENT_AUDIT_SECRET or "sdelat-vychet-consent-audit-v1").encode("utf-8")
    return hmac.new(key, (value or "").encode("utf-8"), hashlib.sha256).hexdigest()


def append_consent_audit(event: dict[str, Any]) -> None:
    """
    Append-only local audit.
    IMPORTANT: production must point CONSENT_AUDIT_PATH to persistent storage
    (or replace with PostgreSQL). Container-local storage alone is not enough.
    """
    if not CONSENT_AUDIT_PATH:
        return
    path = os.path.abspath(CONSENT_AUDIT_PATH)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    line = json.dumps(event, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    with open(path, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def legal_shell(title: str, content: str) -> str:
    address_warning = "" if legal_ready() else """
      <div class="warning"><b>Тестовая конфигурация:</b> адрес оператора ещё не заполнен.
      Перед публичным запуском необходимо задать <code>OPERATOR_ADDRESS</code>.</div>
    """
    return f"""<!doctype html>
<html lang="ru"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{escape(title)} — СделатьВычет</title>
<style>
:root{{--ink:#111;--muted:#6c6c66;--line:#e4e4dd;--paper:#f5f5f0;--accent:#b7ff2a}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--paper);color:var(--ink);font-family:Inter,Arial,sans-serif}}
.wrap{{max-width:880px;margin:0 auto;padding:44px 24px 80px}}.top{{display:flex;justify-content:space-between;gap:20px;align-items:center;margin-bottom:42px}}
.brand{{font-weight:900;letter-spacing:-.04em;font-size:20px}}a{{color:inherit}}.back{{font-size:13px;color:var(--muted)}}
article{{background:white;border:1px solid var(--line);border-radius:18px;padding:38px}}
h1{{font-size:34px;letter-spacing:-.045em;margin:0 0 8px}}h2{{font-size:19px;letter-spacing:-.025em;margin:30px 0 10px}}
p,li{{font-size:14px;line-height:1.65}}li{{margin:6px 0}}.meta{{font-size:12px;color:var(--muted);margin-bottom:28px}}
.note{{padding:14px;background:#f7f7f2;border-radius:10px;margin:14px 0;font-size:13px;line-height:1.6}}
.warning{{padding:14px;background:#fff2de;border:1px solid #f1d3a7;border-radius:10px;margin-bottom:18px;font-size:13px;line-height:1.6}}
.operator{{margin-top:28px;padding-top:20px;border-top:1px solid var(--line);font-size:13px;line-height:1.7}}
code{{background:#eee;padding:2px 5px;border-radius:4px}}@media(max-width:650px){{article{{padding:24px 18px}}h1{{font-size:28px}}}}
</style></head><body><div class="wrap">
<div class="top"><div class="brand">
  <svg class="brandMark" viewBox="0 0 44 44" fill="none" xmlns="http://www.w3.org/2000/svg" aria-label="СделатьВычет">
    <rect x="1" y="1" width="42" height="42" rx="13" fill="#B7FF2A"/>
    <path d="M14 23.5L19.2 28.7L30 16.8" stroke="#142000" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/>
  </svg>
  <div class="brandWord"><b>СделатьВычет</b><span>Налоговый помощник</span></div>
</div><a class="back" href="/">← Вернуться к сервису</a></div>
{address_warning}
<article>{content}</article></div></body></html>"""


def operator_block() -> str:
    return f"""
    <div class="operator">
      <b>Оператор / Исполнитель:</b> {escape(OPERATOR_NAME)}<br>
      Физическое лицо, применяющее специальный налоговый режим «Налог на профессиональный доход»<br>
      ИНН: {escape(OPERATOR_INN)}<br>
      Адрес: {escape(operator_address_display())}<br>
      E-mail: <a href="mailto:{escape(OPERATOR_EMAIL)}">{escape(OPERATOR_EMAIL)}</a>
    </div>
    """


def terms_html() -> str:
    body = f"""
      <h1>Пользовательское соглашение</h1>
      <div class="meta">Версия {TERMS_VERSION}. Действует с 27 августа 2026 года.</div>

      <h2>1. Общие положения</h2>
      <p>Настоящее Пользовательское соглашение регулирует использование сервиса {SERVICE_NAME}.
      Исполнителем является {escape(OPERATOR_NAME)}, физическое лицо, применяющее налоговый режим
      «Налог на профессиональный доход».</p>
      <p>Используя сервис и отдельно принимая настоящее Соглашение, пользователь подтверждает,
      что достиг 18 лет, действует от своего имени и ознакомился с условиями сервиса.</p>

      <h2>2. Что делает СделатьВычет</h2>
      <p>Сервис автоматически анализирует банковскую выписку пользователя, ищет операции,
      которые по описанию могут относиться к расходам, учитываемым при получении налоговых
      вычетов, рассчитывает ориентировочную сумму и формирует персональный информационный отчёт
      и пошаговый маршрут дальнейших действий.</p>
      <div class="note"><b>Важно:</b> результат является предварительным информационным расчётом.
      СделатьВычет не является ФНС России, налоговым органом, адвокатским или аудиторским сервисом
      и не гарантирует предоставление вычета или возврат конкретной суммы.</div>

      <h2>3. Обязанности пользователя</h2>
      <ul>
        <li>загружать собственную банковскую выписку либо документ, который пользователь вправе законно предоставить;</li>
        <li>не загружать медицинские карты, диагнозы, результаты исследований, паспортные сканы и иные документы, не требующиеся для работы сервиса;</li>
        <li>самостоятельно проверять итоговые сведения и документы перед их направлением в ФНС;</li>
        <li>не использовать сервис для противоправных целей или анализа данных третьих лиц без законного основания.</li>
      </ul>

      <h2>4. Стоимость и оказание услуги</h2>
      <p>Бесплатный этап показывает агрегированную оценку найденных расходов и потенциального
      возврата. Доступ к полному персональному отчёту предоставляется после оплаты цены,
      указанной на экране оплаты (на дату этой версии — {REPORT_PRICE_RUB} ₽).</p>
      <p>Услуга по формированию отчёта считается оказанной в момент предоставления пользователю
      доступа к сформированному отчёту. Это положение не ограничивает обязательные права
      потребителя, установленные законодательством Российской Федерации.</p>

      <h2>5. Точность результата</h2>
      <p>Распознавание банковских выписок и классификация операций выполняются автоматически.
      Название продавца, категория банковской операции или факт оплаты сами по себе могут быть
      недостаточны для получения вычета. Окончательное право на вычет определяется законом,
      подтверждающими документами, размером уплаченного НДФЛ и решением налогового органа.</p>

      <h2>6. Персональные данные</h2>
      <p>Обработка персональных данных регулируется отдельной
      <a href="/privacy">Политикой обработки персональных данных</a>.
      Согласие на обработку персональных данных предоставляется пользователем
      <b>отдельно</b> от принятия настоящего Соглашения.</p>

      <h2>7. Интеллектуальные права</h2>
      <p>Программный код, дизайн, структура и материалы сервиса принадлежат правообладателям.
      Пользователю предоставляется право использовать сформированный для него отчёт для личных целей.</p>

      <h2>8. Ответственность и обращения</h2>
      <p>Исполнитель отвечает за нарушение обязательств в пределах, установленных применимым
      законодательством. Никакое положение настоящего Соглашения не исключает ответственность,
      которую нельзя исключить по закону.</p>
      <p>Претензии и запросы направляются на {escape(OPERATOR_EMAIL)}. Исполнитель вправе
      запросить сведения, необходимые для идентификации платежа или конкретного отчёта.</p>

      <h2>9. Изменения</h2>
      <p>Новая редакция Соглашения применяется к действиям, совершённым после её публикации.
      Версия документа фиксируется сервером при загрузке выписки.</p>
      {operator_block()}
    """
    return legal_shell("Пользовательское соглашение", body)


def privacy_html() -> str:
    body = f"""
      <h1>Политика обработки персональных данных</h1>
      <div class="meta">Версия {PRIVACY_VERSION}. Действует с 27 августа 2026 года.</div>

      <h2>1. Оператор</h2>
      <p>Оператором персональных данных при использовании СделатьВычет является
      {escape(OPERATOR_NAME)}. Контакт для обращений: {escape(OPERATOR_EMAIL)}.</p>

      <h2>2. Какие данные обрабатываются</h2>
      <p>В зависимости от содержания банковской выписки сервис технически может получить:</p>
      <ul>
        <li>ФИО пользователя, сведения о банке, счёте или карте, содержащиеся в выписке;</li>
        <li>даты и суммы операций, валюту, описания операций, наименования продавцов и поставщиков услуг;</li>
        <li>агрегированные результаты анализа и выбранные пользователем операции;</li>
        <li>идентификатор анализа, сведения о факте и статусе оплаты;</li>
        <li>техническую запись о принятии документов: дата и время, версии документов, псевдонимизированные хэши IP-адреса и user-agent.</li>
      </ul>
      <p>Платёжные реквизиты банковской карты пользователя СделатьВычет не получает: ввод
      платёжных данных происходит на стороне платёжного провайдера.</p>

      <h2>3. Специальные категории данных</h2>
      <p>СделатьВычет <b>не предназначен для получения диагнозов, медицинских карт, сведений о
      заболеваниях или иных медицинских документов</b>. Сервис анализирует исключительно
      банковские операции в целях классификации расходов для налогового вычета и не ставит
      диагнозов и не делает выводов о состоянии здоровья пользователя.</p>
      <div class="note">Не загружайте в сервис медицинские заключения, рецепты, паспортные
      сканы и другие документы, кроме банковской выписки или поддерживаемой таблицы операций.</div>

      <h2>4. Цели и основания обработки</h2>
      <ul>
        <li><b>анализ выписки и формирование результата</b> — отдельное согласие пользователя и исполнение соглашения с пользователем;</li>
        <li><b>предоставление платного отчёта и обработка платежа</b> — исполнение соглашения;</li>
        <li><b>фиксация факта принятия юридических документов</b> — подтверждение выполнения обязанностей оператора и защита законных прав;</li>
        <li><b>ответы на обращения и выполнение требований закона</b> — исполнение обязанностей, установленных законодательством.</li>
      </ul>

      <h2>5. Как происходит обработка</h2>
      <p>Операции включают сбор, извлечение, систематизацию, использование, временное хранение,
      блокирование и уничтожение данных с использованием средств автоматизации.</p>
      <p><b>Исходный файл банковской выписки не записывается приложением в постоянное хранилище.</b>
      Он передаётся серверу по HTTPS, читается в оперативной памяти и после завершения запроса
      не сохраняется приложением как отдельный файл.</p>
      <p>Результат анализа временно хранится в оперативной памяти приложения до 2 часов и
      удаляется автоматически по истечении этого срока либо при перезапуске процесса.</p>

      <h2>6. Кому могут передаваться данные</h2>
      <p>Для технической работы сервиса используется российская серверная инфраструктура
      хостинг-провайдера. При оплате минимально необходимые сведения о заказе передаются
      платёжному провайдеру. СделатьВычет не передаёт третьим лицам исходную банковскую выписку
      для рекламных целей и не продаёт персональные данные.</p>

      <h2>7. Локализация</h2>
      <p>Первичный сбор и обработка данных пользователей сервиса организуются на серверной
      инфраструктуре, расположенной на территории Российской Федерации. Оператор не
      предусматривает трансграничную передачу банковской выписки.</p>

      <h2>8. Сроки</h2>
      <ul>
        <li>исходный файл — не сохраняется приложением после обработки запроса;</li>
        <li>результат анализа — до 2 часов;</li>
        <li>сведения, необходимые для исполнения требований закона, бухгалтерского/налогового учёта и защиты прав сторон, — в сроки, установленные законодательством;</li>
        <li>аудит факта согласия — в течение срока, необходимого для подтверждения законности обработки и защиты прав оператора.</li>
      </ul>

      <h2>9. Права пользователя</h2>
      <p>Пользователь вправе запросить сведения об обработке своих данных, потребовать их
      уточнения, блокирования или уничтожения при наличии предусмотренных законом оснований,
      а также отозвать согласие. Запрос направляется на {escape(OPERATOR_EMAIL)}.</p>

      <h2>10. Аналитика использования сервиса</h2>
      <p>Для оценки продуктовой воронки сервис ведёт собственную first-party аналитику:
      анонимный идентификатор сессии, UTM-метки, домен-источник и события интерфейса
      (например, выбор файла, успешный анализ, переход к оплате и открытие отчёта).
      В аналитическую базу не записываются названия продавцов, описания банковских
      операций, номера счетов/карт и точные суммы отдельных транзакций.</p>
      <p>Яндекс Метрика подключается только после отдельного разрешения пользователя.
      Вебвизор, карта кликов и отслеживание ссылок отключены. В Метрику передаются только обезличенные просмотры страниц с URL без query-параметров и названия продуктовых целей без содержимого банковской выписки.</p>

      <h2>11. Безопасность и минимизация</h2>
      <p>Оператор применяет принцип минимизации: не хранит исходный PDF после анализа,
      не получает реквизиты платёжной карты, не использует данные выписки для рекламы,
      а технические сетевые идентификаторы в журнале согласий сохраняются в виде хэшей.</p>

      {operator_block()}
    """
    return legal_shell("Политика обработки персональных данных", body)


def consent_html() -> str:
    body = f"""
      <h1>Согласие на обработку персональных данных</h1>
      <div class="meta">Версия {CONSENT_VERSION}. Оформляется отдельно от Пользовательского соглашения.</div>

      <p>Я свободно, своей волей и в своём интересе даю оператору —
      {escape(OPERATOR_NAME)}, адрес: {escape(operator_address_display())},
      согласие на обработку персональных данных в целях анализа банковской выписки,
      поиска потенциальных налоговых вычетов, формирования и предоставления отчёта,
      сопровождения оплаты и исполнения обращений пользователя.</p>

      <h2>1. Перечень данных</h2>
      <p>Согласие распространяется на персональные данные, которые могут содержаться в
      банковской выписке: ФИО, сведения о банке, счёте/карте, даты и суммы операций,
      валюта, описания операций, наименования продавцов/получателей, а также
      технические данные о взаимодействии с сервисом и идентификаторы заказа.</p>

      <h2>2. Действия с данными</h2>
      <p>Разрешаются сбор, извлечение, запись в оперативную память, систематизация,
      использование, временное хранение, блокирование и уничтожение с применением
      средств автоматизации.</p>

      <h2>3. Ограничения</h2>
      <p>Согласие <b>не является согласием на распространение</b> персональных данных.
      Оператор не просит пользователя предоставлять диагнозы, медицинские карты,
      медицинские заключения, биометрические данные или паспортные сканы.</p>

      <h2>4. Срок</h2>
      <p>Согласие действует с момента его предоставления до достижения целей обработки
      или до его отзыва, если у оператора отсутствует иное законное основание для
      продолжения обработки. Исходный файл выписки приложением после анализа не
      сохраняется; результат анализа хранится до 2 часов.</p>

      <h2>5. Отзыв</h2>
      <p>Согласие может быть отозвано обращением на
      <a href="mailto:{escape(OPERATOR_EMAIL)}">{escape(OPERATOR_EMAIL)}</a>.
      В обращении следует указать сведения, позволяющие установить, к какой обработке
      относится требование.</p>

      <div class="note"><b>Отдельное действие:</b> на странице загрузки это Согласие
      подтверждается отдельным чекбоксом и не объединено с принятием Пользовательского соглашения.</div>

      {operator_block()}
    """
    return legal_shell("Согласие на обработку персональных данных", body)


def cleanup_analyses() -> None:
    now = time.time()
    expired = [k for k, v in ANALYSES.items() if now - v.get("created_at", now) > ANALYSIS_TTL_SECONDS]
    for k in expired:
        ANALYSES.pop(k, None)


def get_analysis_or_404(analysis_id: str) -> dict[str, Any]:
    cleanup_analyses()
    item = ANALYSES.get(analysis_id)
    if not item:
        raise HTTPException(404, "Результат анализа не найден или срок хранения истёк. Загрузите выписку ещё раз.")
    return item


def yookassa_request(method: str, path: str, payload: dict[str, Any] | None = None, idempotence_key: str | None = None) -> dict[str, Any]:
    if not (YOOKASSA_SHOP_ID and YOOKASSA_SECRET_KEY):
        raise HTTPException(503, "Оплата пока не подключена: добавьте YOOKASSA_SHOP_ID и YOOKASSA_SECRET_KEY в Timeweb.")
    url = "https://api.yookassa.ru/v3" + path
    token = base64.b64encode(f"{YOOKASSA_SHOP_ID}:{YOOKASSA_SECRET_KEY}".encode("utf-8")).decode("ascii")
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
    headers = {
        "Authorization": f"Basic {token}",
        "Content-Type": "application/json",
    }
    if idempotence_key:
        headers["Idempotence-Key"] = idempotence_key
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise HTTPException(502, f"ЮKassa вернула ошибку: {detail[:500]}")
    except Exception as e:
        raise HTTPException(502, f"Не удалось связаться с ЮKassa: {e}")


def public_base_url(request: Request) -> str:
    if PUBLIC_BASE_URL:
        return PUBLIC_BASE_URL
    return str(request.base_url).rstrip("/")


# Recognition signatures for major Russian retail banks.
# Detection != guaranteed bank-specific parsing: unsupported layouts fall through
# to the strict universal deduction-oriented parser.
BANK_SIGNATURES = [
    ("sber", "Сбер", ["ПАО СБЕРБАНК", "СБЕРБАНК", "SBERBANK.RU", "СБЕР БАНК"]),
    ("vtb", "ВТБ", ["БАНК ВТБ", "ВТБ (ПАО)", "VTB.RU", "ПАО ВТБ"]),
    ("gazprombank", "Газпромбанк", ["ГАЗПРОМБАНК", "БАНК ГПБ", "GAZPROMBANK"]),
    ("alfa", "Альфа-Банк", ["АЛЬФА-БАНК", "АЛЬФА БАНК", "ALFABANK", "ALFA-BANK"]),
    ("tbank", "Т-Банк", ["АО «ТБАНК»", 'АО "ТБАНК"', "Т-БАНК", "TBANK.RU", "ТИНЬКОФФ БАНК", "TINKOFF BANK"]),
    ("rshb", "Россельхозбанк", ["РОССЕЛЬХОЗБАНК", "RSHB.RU"]),
    ("mkb", "МКБ", ["МОСКОВСКИЙ КРЕДИТНЫЙ БАНК", "МКБ", "MKB.RU"]),
    ("domrf", "Банк ДОМ.РФ", ["БАНК ДОМ.РФ", "DOM.RF BANK", "ДОМ.РФ БАНК"]),
    ("sovcombank", "Совкомбанк", ["СОВКОМБАНК", "ХАЛВА", "SOVCOMBANK"]),
    ("raiffeisen", "Райффайзенбанк", ["РАЙФФАЙЗЕНБАНК", "RAIFFEISENBANK", "RAIFFEISEN BANK"]),
    ("bmbank", "БМ-Банк", ["БМ-БАНК", "БМ БАНК"]),
    ("novikom", "Новикомбанк", ["НОВИКОМБАНК"]),
    ("bankrossiya", "Банк Россия", ['БАНК "РОССИЯ"', "АБ РОССИЯ"]),
    ("bspb", "Банк Санкт-Петербург", ["БАНК САНКТ-ПЕТЕРБУРГ", "BSPB.RU"]),
    ("akbars", "Ак Барс Банк", ["АК БАРС", "AK BARS"]),
    ("rencap", "РенКап Банк", ["РЕНКАП БАНК", "РЕНЕССАНС КРЕДИТ", "RENCAP"]),
    ("otp", "ОТП Банк", ["ОТП БАНК", "OTP BANK"]),
    ("ozon", "Озон Банк", ["ОЗОН БАНК", "OZON BANK"]),
    ("uralsib", "Уралсиб", ["УРАЛСИБ", "URALSIB"]),
    ("mts", "МТС Банк", ["МТС БАНК", "MTS BANK"]),
    ("unicredit", "ЮниКредит Банк", ["ЮНИКРЕДИТ БАНК", "UNICREDIT BANK"]),
    ("tkb", "ТКБ Банк", ["ТКБ БАНК", "TRANSKAPITALBANK"]),
    ("yandex", "Яндекс Банк", ["ЯНДЕКС БАНК", "YANDEX BANK"]),
    ("atb", "Азиатско-Тихоокеанский Банк", ["АЗИАТСКО-ТИХООКЕАНСКИЙ БАНК", "АТБ БАНК"]),
    ("expobank", "Экспобанк", ["ЭКСПОБАНК", "EXPOBANK"]),
    ("tochka", "Точка", ["БАНК ТОЧКА", "ТОЧКА БАНК"]),
    ("ubrir", "УБРиР", ["УБРИР", "УРАЛЬСКИЙ БАНК РЕКОНСТРУКЦИИ И РАЗВИТИЯ"]),
    ("russianstandard", "Русский Стандарт", ["РУССКИЙ СТАНДАРТ", "RUSSIAN STANDARD"]),
    ("credit_europe", "Кредит Европа Банк", ["КРЕДИТ ЕВРОПА БАНК", "CREDIT EUROPE BANK"]),
    ("loko", "Локо-Банк", ["ЛОКО-БАНК", "LOKO BANK"]),
]


def recognize_bank(text: str) -> tuple[str, str]:
    upper = text.upper()
    for key, display, markers in BANK_SIGNATURES:
        if any(marker.upper() in upper for marker in markers):
            return key, display
    return "unknown", "Неизвестный банк"


CATEGORY_META = {
    "medicine": {"name": "Медицина", "emoji": "🏥", "confidence": "high", "note": "Похоже на оплату медицинских услуг. Для вычета потребуется подтверждение от медицинской организации."},
    "pharmacy": {"name": "Аптеки / лекарства", "emoji": "💊", "confidence": "verify", "note": "Аптечная покупка сама по себе не гарантирует вычет: обычно нужны назначение врача и подтверждающие документы."},
    "fitness": {"name": "Спорт / фитнес", "emoji": "🏋️", "confidence": "verify", "note": "Нужно проверить, дает ли организация право на спортивный вычет за соответствующий год."},
    "education": {"name": "Обучение", "emoji": "🎓", "confidence": "verify", "note": "Нужно подтвердить, что платеж относится к обучению и организация/ИП соответствует условиям вычета."},
    "insurance": {"name": "Страхование", "emoji": "🛡️", "confidence": "verify", "note": "Нужно определить тип договора. ОСАГО/каско не являются обычным социальным вычетом."},
    "donation": {"name": "Благотворительность", "emoji": "❤️", "confidence": "verify", "note": "Нужно проверить получателя и документы. Такие расходы не включаются в базовый расчет автоматически."},
}

RULES = [
    ("medicine", [r"MCC(?:8011|8062|8099|8021|8031|8041|8042|8049)\b", r"MEDSI(?:\b|_)", r"DERAJS", r"DERAYS", r"ДЭРАЙС", r"\bMEDSKAN\w*\b", r"\bKLINIKA\b", r"КЛИНИК", r"MEDICINSK", r"МЕДИЦ", r"STOMAT", r"СТОМАТ", r"\bDENT(?:AL)?\b", r"КОСМЕТОЛ", r"\bBESTCLIN"]),
    ("pharmacy", [r"MCC5912\b", r"\bAPTEKA\b", r"АПТЕК", r"АПТЕЧ", r"\bGORZDRAV\b", r"\bMSKAPT"]),
    ("fitness", [r"MCC7997\b", r"\bSTROYTELO\b", r"\bFITNESS\b", r"\bFITNES\b", r"ФИТНЕС", r"\bEMS\b"]),
    ("education", [r"MCC(?:8220|8241|8244|8249|8351)\b", r"\bSKILLBOX\b", r"\bNETOLOGY\b", r"GEEKBRAINS", r"\bUNIVERS", r"УНИВЕРСИТ", r"\bSCHOOL\b", r"ШКОЛ", r"\bCOURSE\b", r"КУРС"]),
    ("insurance", [r"ИНГОССТРАХ", r"\bINGOS", r"\bINSURANCE\b", r"СТРАХОВ", r"MCC6300\b"]),
    ("donation", [r"БЛАГОТВОР", r"\bCHARITY\b", r"MCC8398\b"]),
]

TX_RE = re.compile(r"(?ms)^(\d{2}\.\d{2}\.\d{4})\s+([A-ZА-Я0-9_]+)\s+(.*?)(?=^\d{2}\.\d{2}\.\d{4}\s+[A-ZА-Я0-9_]+\s+|\Z)")
RUR_AMOUNT_RE = re.compile(r"(?<!\d)([-+]?\d[\d ]*[,.]\d{2})\s+RUR")


def money_to_float(s: str) -> float:
    return float(s.replace("\xa0", "").replace(" ", "").replace(",", "."))


def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


def normalize_date(v: Any) -> str:
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    s = str(v or "").strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%d.%m.%Y")
        except Exception:
            pass
    return s[:10]


def extract_merchant(desc: str) -> str:
    m = re.search(r"место совершения операции:\s*(.*?)\s+MCC\d{4}\b", desc, re.I)
    if m:
        place = m.group(1).strip()
        parts = [p.strip() for p in place.split("\\") if p.strip()]
        if parts:
            merchant = re.sub(r"^\d{5,}\s*", "", parts[-1]).strip()
            return merchant[:100]
    m = re.search(r"\bв\s+(.+?)\s+через\s+Систему быстрых платежей", desc, re.I)
    if m:
        return m.group(1).strip()[:100]
    if "Оплата страхового продукта" in desc:
        return "Страховой продукт"
    return "Не удалось определить"


TBANK_PAIR_RE = re.compile(
    r"(?m)^(\d{2}\.\d{2}\.\d{4})\n(\d{2}:\d{2})\n"
    r"(\d{2}\.\d{2}\.\d{4})\n(\d{2}:\d{2})\n"
)


def extract_tbank_merchant(description: str) -> str:
    s = normalize(description)

    m = re.match(r"Оплата в\s+(.+)", s, re.I)
    if m:
        merchant = m.group(1).strip()
        location_patterns = [
            r"\s+Gorod\s+Moskva\s+RUS$",
            r"\s+MOSKVA\s+RUS$",
            r"\s+Moskva\s+RUS$",
            r"\s+MOSCOW\s+RUS$",
            r"\s+Moscow\s+RU$",
            r"\s+Moskva\s+RU$",
            r"\s+G\s+MOSKVA\s+RU$",
            r"\s+Москва\s+Россия$",
            r"\s+[A-Za-zА-Яа-яЁё.\-]+\s+(?:RUS|RU|BLR)$",
        ]
        for pattern in location_patterns:
            merchant = re.sub(pattern, "", merchant, flags=re.I).strip()
        return merchant[:100] or "Оплата картой"

    m = re.match(r"Оплата услуг\s+(.+)", s, re.I)
    if m:
        return m.group(1).strip()[:100]

    return s[:100] or "Операция"


def parse_tbank_pdf_text(text: str) -> list[dict[str, Any]]:
    pairs = list(TBANK_PAIR_RE.finditer(text))
    if not pairs:
        raise ValueError("Не удалось найти операции в выписке Т-Банка.")

    txs: list[dict[str, Any]] = []

    for i, pair in enumerate(pairs):
        end = pairs[i + 1].start() if i + 1 < len(pairs) else len(text)
        block = text[pair.end():end]

        cut_positions = []
        for marker in (
            'АО «ТБанк»',
            'АО "ТБанк"',
            "АКЦИОНЕРНОЕ ОБЩЕСТВО «ТБАНК»",
            "Дата и время\nоперации",
            "Пополнения:",
            "Расходы:",
            "С уважением,",
        ):
            pos = block.find(marker)
            if pos >= 0:
                cut_positions.append(pos)
        if cut_positions:
            block = block[:min(cut_positions)]

        m = re.match(
            r"(?s)^\s*"
            r"([+-]?\d[\d ]*[.,]\d{2})\s+(\S+)\s+"
            r"([+-]?\d[\d ]*[.,]\d{2})\s+₽\s+"
            r"(.*?)\s*$",
            block,
        )
        if not m:
            continue

        operation_amount, operation_currency, card_amount, description = m.groups()

        card_match = re.search(r"(?:^|\s)(—|\d{4})\s*$", description)
        card_last4 = ""
        if card_match:
            card_last4 = card_match.group(1)
            description = description[:card_match.start()].strip()

        desc = normalize(description)

        txs.append({
            "date": pair.group(1),
            "code": f"TBANK_{i + 1:05d}",
            "description": desc,
            "merchant": extract_tbank_merchant(desc),
            "amount": money_to_float(card_amount),
            "bank": "Т-Банк",
            "posting_date": pair.group(3),
            "card_last4": card_last4,
            "operation_currency": operation_currency,
            "operation_amount": money_to_float(operation_amount),
        })

    if not txs:
        raise ValueError("Не удалось распознать операции в выписке Т-Банка.")
    return txs



SBER_PAIR_RE = re.compile(
    r"(?m)^(\d{2}\.\d{2}\.\d{4})\n"
    r"(\d{2}\.\d{2}\.\d{4})\n"
    r"(\d{2}:\d{2})\n"
    r"(\d{6})\n"
)

SBER_AMOUNT_LINE_RE = re.compile(
    r"^\s*(\+?\d[\d\u00a0 ]*,\d{2})(?:\s+.*)?$"
)


def extract_sber_merchant(description: str) -> str:
    s = normalize(description)

    # Strip the standard card suffix.
    s = re.sub(r"\s*Операция по карте\s+\*+\d{4}\s*$", "", s, flags=re.I).strip()

    # Strip standard SBP boilerplate.
    s = re.sub(
        r"\s*Покупка по СБП в ТСТ (?:другого банка|Сбербанка)\.*\s*$",
        "",
        s,
        flags=re.I,
    ).strip()

    # Strip common location prefixes from merchant descriptors.
    s = re.sub(
        r"^(?:MOSCOW|MOSKVA|SANKT-PETERBU|Krasnogorsk)\s+",
        "",
        s,
        flags=re.I,
    ).strip()

    # QR suffixes are noise for merchant presentation, but keeping merchant stem is useful.
    s = re.sub(r"_(?:P|E)_QR\.?$", "", s, flags=re.I).strip()

    return s[:100] or "Операция"


def parse_sber_pdf_text(text: str) -> list[dict[str, Any]]:
    """
    Parse Sber's 'Выписка по счёту кредитной карты'.

    Stable block structure:
      operation date
      posting date
      time
      6-digit authorization code
      bank category
      merchant/description
      amount in RUB [and balance]
      optional foreign-currency amount
      balance
    """
    pairs = list(SBER_PAIR_RE.finditer(text))
    if not pairs:
        raise ValueError("Не удалось найти операции в выписке Сбера.")

    txs: list[dict[str, Any]] = []

    for i, pair in enumerate(pairs):
        end = pairs[i + 1].start() if i + 1 < len(pairs) else len(text)
        block = text[pair.end():end]

        # Remove repeated page headers/footers and document footer.
        for marker in (
            "Продолжение на следующей странице",
            "Дата формирования документа",
            "Выписка по счёту кредитной карты Страница",
        ):
            pos = block.find(marker)
            if pos >= 0:
                block = block[:pos]

        lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
        if len(lines) < 2:
            continue

        bank_category = lines[0]
        amount_idx = None
        raw_amount = None

        # First amount-looking line after category is the transaction RUB amount.
        # The second number on the same line, if present, is the running balance.
        for j, line in enumerate(lines[1:], start=1):
            m = SBER_AMOUNT_LINE_RE.match(line)
            if m:
                raw_amount = m.group(1)
                amount_idx = j
                break

        if amount_idx is None or raw_amount is None:
            continue

        description = normalize(" ".join(lines[1:amount_idx]))
        amount_abs = money_to_float(raw_amount.replace("+", ""))

        category_lower = bank_category.lower()
        desc_lower = description.lower()

        is_credit = (
            raw_amount.startswith("+")
            or "возврат" in category_lower
            or "перевод на карту" in category_lower
            or "пополнение" in category_lower
            or "зачислен" in category_lower
            or "возврат" in desc_lower
        )
        amount = amount_abs if is_credit else -amount_abs

        card_match = re.search(r"\*+(\d{4})", description)
        card_last4 = card_match.group(1) if card_match else ""

        txs.append({
            "date": pair.group(1),
            "code": f"SBER_{pair.group(4)}_{i + 1:04d}",
            # Include Sber's category in searchable description but keep merchant separately.
            "description": f"{bank_category} {description}",
            "merchant": extract_sber_merchant(description),
            "amount": amount,
            "bank": "Сбер",
            "posting_date": pair.group(2),
            "auth_code": pair.group(4),
            "card_last4": card_last4,
            "bank_category": bank_category,
        })

    if not txs:
        raise ValueError("Не удалось распознать операции в выписке Сбера.")
    return txs



VTB_ROW_START_RE = re.compile(
    r"(?m)^(\d{2}\.\d{2}\.\d{4})(?:\s+|\n)(\d{2}:\d{2}(?::\d{2})?)"
)

RAIFF_ROW_RE = re.compile(
    r"(?m)^\s*(\d+)\s+(\d{2}\.\d{2}\.\d{2,4})\s+(\d{2}\.\d{2}\.\d{2,4})\s+"
)

GENERIC_DATE_RE = re.compile(
    r"(?m)^(?:\d+\s+)?(\d{2}[./-]\d{2}[./-]\d{2,4})(?:\s+(\d{2}:\d{2}(?::\d{2})?))?"
)
GENERIC_MONEY_RE = re.compile(
    r"(?<!\d)([+-]?(?:\d{1,3}(?:[ \u00a0.,]\d{3})+|\d+)(?:[.,-]\d{2}))\s*(₽|RUB|RUR)?",
    re.I,
)

DEBIT_WORDS = (
    "ОПЛАТА", "ПОКУПКА", "СПИСАН", "ПЕРЕВОД НА", "ПЕРЕВОД ПО", "СНЯТИ",
    "ВЫДАЧА НАЛИЧ", "КОМИСС", "ПЛАТА", "PAYMENT", "PURCHASE", "DEBIT",
)
CREDIT_WORDS = (
    "ПОПОЛН", "ЗАЧИСЛ", "ВОЗВРАТ", "ПЕРЕВОД С НОМЕРА", "ПЕРЕВОД ОТ",
    "ВНЕСЕНИЕ НАЛИЧ", "CREDIT",
)


def _parse_money_generic(raw: str) -> float:
    s = raw.replace("\u00a0", " ").strip()
    # Raiffeisen old-style "2,000-00" -> 2000.00
    if re.match(r"^-?\d{1,3}(?:,\d{3})*-\d{2}$", s):
        neg = s.startswith("-")
        s2 = s.lstrip("-").replace(",", "").replace("-", ".")
        val = float(s2)
        return -val if neg else val

    # Russian spaces thousands + comma decimal.
    if "," in s and "." not in s:
        s = s.replace(" ", "").replace(",", ".")
    else:
        # If both separators occur, assume commas/spaces are grouping.
        s = s.replace(" ", "").replace(",", "")
    return float(s)


def extract_vtb_merchant(desc: str) -> str:
    s = normalize(desc)
    # Common VTB descriptions put the useful counterparty/merchant at the end.
    for marker in ("ОПИСАНИЕ ОПЕРАЦИИ", "ОПИСАНИЕ:"):
        s = s.replace(marker, " ")
    s = re.sub(r"\b(?:RUS|RUB|RUВ)\b", " ", s, flags=re.I)
    s = normalize(s)
    return s[-100:] if len(s) > 100 else (s or "Операция ВТБ")


def parse_vtb_pdf_text(text: str) -> list[dict[str, Any]]:
    """
    Supports the common VTB retail statement family described with:
    operation datetime, bank processing date, operation amount, card/account
    income/expense and description.
    """
    starts = list(VTB_ROW_START_RE.finditer(text))
    if not starts:
        raise ValueError("Не удалось найти операции в выписке ВТБ.")

    txs = []
    for i, m in enumerate(starts):
        end = starts[i + 1].start() if i + 1 < len(starts) else len(text)
        block = normalize(text[m.end():end])
        if not block:
            continue

        # Remove processing dates/times before looking for monetary values.
        # Otherwise "08.06.2023" can be misread as 8.06.
        money_block = re.sub(r"\b\d{2}\.\d{2}\.\d{4}\b", " ", block)
        money_block = re.sub(r"\b\d{2}:\d{2}(?::\d{2})?\b", " ", money_block)

        amounts = []
        for mm in GENERIC_MONEY_RE.finditer(money_block):
            try:
                amounts.append((_parse_money_generic(mm.group(1)), mm.group(0), mm.start()))
            except Exception:
                pass
        if not amounts:
            continue

        upper = block.upper()
        # Prefer explicit "расход" value if textual labels survived extraction.
        expense_match = re.search(r"(?:РАСХОД|ДЕБЕТ)\D{0,20}(\d[\d \u00a0.,]*[.,]\d{2})", block, re.I)
        income_match = re.search(r"(?:ПРИХОД|КРЕДИТ)\D{0,20}(\d[\d \u00a0.,]*[.,]\d{2})", block, re.I)

        if expense_match:
            amount = -abs(_parse_money_generic(expense_match.group(1)))
        elif income_match and not any(w in upper for w in DEBIT_WORDS):
            amount = abs(_parse_money_generic(income_match.group(1)))
        else:
            # Signed operation amount if present.
            signed = [a for a in amounts if a[0] < 0]
            if signed:
                amount = signed[0][0]
            else:
                # Purchase/payment descriptions are debits; transfers-in/refunds credits.
                op_amount = abs(amounts[0][0])
                if any(w in upper for w in CREDIT_WORDS) and not any(w in upper for w in DEBIT_WORDS):
                    amount = op_amount
                else:
                    amount = -op_amount

        # Remove numbers to produce a cleaner merchant-like description.
        desc = re.sub(r"\d[\d \u00a0.,-]*\s*(?:RUB|RUR|RUВ|₽)?", " ", block, flags=re.I)
        desc = normalize(desc)

        txs.append({
            "date": m.group(1),
            "code": f"VTB_{i+1:05d}",
            "description": desc,
            "merchant": extract_vtb_merchant(desc),
            "amount": round(amount, 2),
            "bank": "ВТБ",
            "parser_confidence": "bank_specific_beta",
        })
    if not txs:
        raise ValueError("Не удалось распознать операции в выписке ВТБ.")
    return txs


def extract_raiff_merchant(desc: str) -> str:
    s = normalize(desc)
    # Card lines commonly: CARD **9997 11AUG RUR 443 MERCHANT CITY
    m = re.search(r"\bCARD\b.*?\b(?:RUR|RUB)\b\s+[\d.]+\s+(.+)$", s, re.I)
    if m:
        merchant = m.group(1)
        merchant = re.sub(r"\s+(?:MOSKVA|MOSCOW|SANKT-PETERBU|RUS|RU)$", "", merchant, flags=re.I)
        return merchant[:100].strip()
    # SBP payment lines: "Оплата покупки 329.00 RUB МПП - Платные дороги."
    m = re.search(r"Оплата покупки(?: по карте)?[.\s]+(?:[\d.,]+\s+RUB\s+)?(.+)", s, re.I)
    if m:
        return m.group(1)[:100].strip()
    return s[-100:] if len(s) > 100 else (s or "Операция Райффайзен")


def parse_raiffeisen_pdf_text(text: str) -> list[dict[str, Any]]:
    rows = list(RAIFF_ROW_RE.finditer(text))
    if not rows:
        raise ValueError("Не удалось найти операции в выписке Райффайзенбанка.")
    txs = []
    for i, m in enumerate(rows):
        end = rows[i+1].start() if i+1 < len(rows) else len(text)
        block = normalize(text[m.end():end])
        upper = block.upper()

        # Raiff account statements usually carry an amount like 2,000-00.
        candidates = []
        for mm in re.finditer(r"(?<!\d)(\d{1,3}(?:,\d{3})*-\d{2}|\d[\d \u00a0]*[.,]\d{2})(?!\d)", block):
            try:
                val = _parse_money_generic(mm.group(1))
                candidates.append((val, mm.start(), mm.group(1)))
            except Exception:
                pass
        if not candidates:
            continue

        # Skip long account/correspondent numbers by requiring reasonable transaction magnitude token.
        amount_abs = abs(candidates[0][0])
        if amount_abs == 0:
            continue

        if any(w in upper for w in CREDIT_WORDS) and not any(w in upper for w in DEBIT_WORDS):
            amount = amount_abs
        else:
            amount = -amount_abs

        desc = block
        txs.append({
            "date": normalize_date(m.group(2)),
            "code": f"RAIFF_{m.group(1)}",
            "description": desc,
            "merchant": extract_raiff_merchant(desc),
            "amount": round(amount, 2),
            "bank": "Райффайзенбанк",
            "parser_confidence": "bank_specific_beta",
        })
    if not txs:
        raise ValueError("Не удалось распознать операции в выписке Райффайзенбанка.")
    return txs


def _tax_signal_in_text(s: str) -> bool:
    up = s.upper()
    for _cat, pats in RULES:
        if any(re.search(p, up, re.I) for p in pats):
            return True
    return False


def parse_generic_tax_pdf_text(text: str, bank_display: str = "Банк") -> list[dict[str, Any]]:
    """
    Strict fallback for digital PDFs from unsupported banks.

    It does NOT try to reconstruct the whole ledger. It only emits transaction
    blocks that contain a tax-deduction signal AND where an amount can be
    associated with the block with reasonable confidence.
    """
    starts = list(GENERIC_DATE_RE.finditer(text))
    if not starts:
        raise ValueError("В PDF не удалось найти таблицу операций.")

    out = []
    seen = set()
    for i, m in enumerate(starts):
        end = starts[i+1].start() if i+1 < len(starts) else min(len(text), m.end()+1200)
        block_raw = text[m.start():end]
        block = normalize(block_raw)
        if not _tax_signal_in_text(block):
            continue

        upper = block.upper()
        # Avoid document headers that happen to mention medical/insurance words.
        if len(block) > 900:
            continue

        money_scan_block = re.sub(r"\b\d{2}[./-]\d{2}[./-]\d{2,4}\b", " ", block)
        money_scan_block = re.sub(r"\b\d{2}:\d{2}(?::\d{2})?\b", " ", money_scan_block)

        money = []
        for mm in GENERIC_MONEY_RE.finditer(money_scan_block):
            raw = mm.group(1)
            try:
                val = _parse_money_generic(raw)
            except Exception:
                continue
            # Reject obvious years/account fragments and zeroes.
            if val == 0 or abs(val) > 10_000_000:
                continue
            score = 0
            if mm.group(2):
                score += 3
            if raw.startswith(("+", "-")):
                score += 2
            # Transaction amount usually sits near purchase/merchant text.
            dist = min(
                [abs(mm.start()-p) for p in [upper.find("ОПЛАТ"), upper.find("ПОКУП"), upper.find("MED"), upper.find("АПТЕК"), upper.find("FIT") ] if p >= 0]
                or [999]
            )
            if dist < 250:
                score += 2
            money.append((score, mm.start(), val, raw))

        if not money:
            continue
        money.sort(key=lambda x: (-x[0], x[1]))
        best = money[0]
        # Require some structural evidence; otherwise do not invent a number.
        if best[0] < 2:
            continue

        val = abs(best[2])
        if any(w in upper for w in CREDIT_WORDS) and not any(w in upper for w in DEBIT_WORDS):
            amount = val
        else:
            amount = -val

        fingerprint = (m.group(1), round(amount, 2), block[:160])
        if fingerprint in seen:
            continue
        seen.add(fingerprint)

        # Use the tax-relevant part of the block as merchant-ish label.
        merchant = block[:100]
        for token in ("MEDSI", "MEDSKAN", "АПТЕК", "FITNESS", "ФИТНЕС", "СТРАХ", "INSURANCE", "КЛИНИК", "STOMAT"):
            pos = upper.find(token)
            if pos >= 0:
                merchant = block[max(0, pos-30):pos+70]
                break

        out.append({
            "date": normalize_date(m.group(1).replace("/", ".").replace("-", ".")),
            "code": f"GENERIC_{i+1:05d}",
            "description": block,
            "merchant": normalize(merchant),
            "amount": round(amount, 2),
            "bank": bank_display,
            "parser_confidence": "strict_generic",
        })

    # Returning no candidates is valid: statement may simply have no deductible spend.
    return out


def parse_alfa_pdf_text(text: str) -> list[dict[str, Any]]:
    txs = []
    for m in TX_RE.finditer(text):
        date, code, body = m.groups()
        amounts = RUR_AMOUNT_RE.findall(body)
        if not amounts:
            continue
        amount = money_to_float(amounts[-1])
        desc = normalize(body)
        txs.append({
            "date": date,
            "code": code,
            "description": desc,
            "merchant": extract_merchant(desc),
            "amount": amount,
            "bank": "Альфа-Банк",
        })
    if not txs:
        raise ValueError("Не удалось распознать операции в выписке Альфа-Банка.")
    return txs


def extract_pdf_text(data: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    return "\n".join((p.extract_text() or "") for p in reader.pages)


def detect_pdf_bank_from_text(text: str) -> str:
    upper = text.upper()
    key, _display = recognize_bank(text)

    # Exact known layouts first.
    if key == "sber" and (
        "ВЫПИСКА ПО СЧЁТУ КРЕДИТНОЙ КАРТЫ" in upper
        or "РАСШИФРОВКА ОПЕРАЦИЙ" in upper
        or SBER_PAIR_RE.search(text)
    ):
        return "sber"

    if key == "tbank" and (
        "СПРАВКА О ДВИЖЕНИИ СРЕДСТВ" in upper
        or TBANK_PAIR_RE.search(text)
    ):
        return "tbank"

    if key == "alfa" or ("АЛЬФА" in upper and RUR_AMOUNT_RE.search(text)):
        return "alfa"

    if key == "vtb":
        return "vtb"

    if key == "raiffeisen":
        return "raiffeisen"

    if key != "unknown":
        return f"known:{key}"

    return "unknown"



def parse_alfa_pdf(data: bytes) -> list[dict[str, Any]]:
    return parse_alfa_pdf_text(extract_pdf_text(data))


def parse_tbank_pdf(data: bytes) -> list[dict[str, Any]]:
    return parse_tbank_pdf_text(extract_pdf_text(data))


def parse_sber_pdf(data: bytes) -> list[dict[str, Any]]:
    return parse_sber_pdf_text(extract_pdf_text(data))

def _guess_column(headers: list[str], variants: list[str]) -> str | None:
    low = {h.lower().strip(): h for h in headers}
    for v in variants:
        for lk, original in low.items():
            if v in lk:
                return original
    return None


def parse_csv(data: bytes) -> list[dict[str, Any]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            pass
    if text is None:
        raise ValueError("Не удалось определить кодировку CSV.")
    try:
        delimiter = csv.Sniffer().sniff(text[:5000], delimiters=",;\t").delimiter
    except Exception:
        delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = reader.fieldnames or []
    date_col = _guess_column(headers, ["дата", "date"])
    desc_col = _guess_column(headers, ["опис", "назнач", "merchant", "операц", "description"])
    amount_col = _guess_column(headers, ["сумм", "amount"])
    if not (date_col and desc_col and amount_col):
        raise ValueError("В CSV нужны колонки с датой, описанием операции и суммой.")
    out = []
    for i, row in enumerate(reader):
        raw_amt = str(row.get(amount_col, "")).replace("RUR", "").replace("₽", "").strip()
        try:
            amount = money_to_float(raw_amt)
        except Exception:
            continue
        desc = normalize(str(row.get(desc_col, "")))
        out.append({"date": normalize_date(row.get(date_col, "")), "code": f"CSV_{i+1}", "description": desc, "merchant": desc[:100] or "Операция", "amount": amount})
    return out


def parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Пустой XLSX.")
    headers = [normalize(str(x or "")) for x in rows[0]]
    date_col = _guess_column(headers, ["дата", "date"])
    desc_col = _guess_column(headers, ["опис", "назнач", "merchant", "операц", "description"])
    amount_col = _guess_column(headers, ["сумм", "amount"])
    if not (date_col and desc_col and amount_col):
        raise ValueError("В XLSX нужны колонки с датой, описанием операции и суммой.")
    idx = {h: n for n, h in enumerate(headers)}
    out = []
    for i, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        amount_val = row[idx[amount_col]] if idx[amount_col] < len(row) else None
        try:
            amount = float(amount_val) if isinstance(amount_val, (int, float)) else money_to_float(str(amount_val or ""))
        except Exception:
            continue
        date_val = row[idx[date_col]] if idx[date_col] < len(row) else ""
        desc_val = row[idx[desc_col]] if idx[desc_col] < len(row) else ""
        desc = normalize(str(desc_val or ""))
        out.append({"date": normalize_date(date_val), "code": f"XLSX_{i}", "description": desc, "merchant": desc[:100] or "Операция", "amount": amount})
    return out


def classify(tx: dict[str, Any]) -> dict[str, Any] | None:
    if tx["amount"] >= 0:
        return None
    s = tx["description"].upper()
    for category, patterns in RULES:
        if any(re.search(p, s, re.I) for p in patterns):
            meta = CATEGORY_META[category]
            fingerprint = hashlib.sha1(f'{tx["date"]}|{tx["code"]}|{tx["amount"]}|{tx["description"]}'.encode("utf-8")).hexdigest()[:14]
            return {
                "id": fingerprint,
                "date": tx["date"],
                "year": tx["date"][-4:] if len(tx["date"]) >= 4 else "",
                "merchant": tx["merchant"],
                "amount": round(abs(tx["amount"]), 2),
                "category": category,
                "category_name": meta["name"],
                "emoji": meta["emoji"],
                "confidence": meta["confidence"],
                "status": "Похоже подходит" if meta["confidence"] == "high" else "Нужно подтвердить",
                "note": meta["note"],
                "selected": True,
            }
    return None


def calculate_refund(candidates: list[dict[str, Any]], selected_ids: set[str] | None = None) -> dict[str, Any]:
    # Консервативная модель MVP: 13%, общий социальный лимит 150 000 ₽ на каждый год.
    by_year = defaultdict(float)
    selected = []
    for c in candidates:
        use = c["selected"] if selected_ids is None else c["id"] in selected_ids
        if use:
            selected.append(c)
            by_year[c["year"]] += c["amount"]
    capped = {y: min(v, 150_000.0) for y, v in by_year.items()}
    base = sum(capped.values())
    return {
        "selected_amount": round(sum(c["amount"] for c in selected), 2),
        "tax_base_after_conservative_cap": round(base, 2),
        "refund_from": round(base * 0.13, 2),
        "by_year": [{"year": y, "selected_amount": round(by_year[y], 2), "tax_base": round(capped[y], 2), "refund_from": round(capped[y] * 0.13, 2)} for y in sorted(by_year)],
    }


def analyze_transactions(txs: list[dict[str, Any]]) -> dict[str, Any]:
    candidates = [c for tx in txs if (c := classify(tx)) is not None]
    base = calculate_refund(candidates)
    all_potential = calculate_refund(candidates, {c["id"] for c in candidates})
    group = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for c in candidates:
        group[c["category"]]["count"] += 1
        group[c["category"]]["amount"] += c["amount"]
    groups = []
    for k, v in group.items():
        meta = CATEGORY_META[k]
        groups.append({"category": k, "name": meta["name"], "emoji": meta["emoji"], "count": v["count"], "amount": round(v["amount"], 2), "confidence": meta["confidence"]})
    groups.sort(key=lambda x: x["amount"], reverse=True)
    fast_candidates = [c for c in candidates if c["category"] in {"medicine", "fitness", "education", "insurance"}]
    fast_path = calculate_refund(fast_candidates, {c["id"] for c in fast_candidates})
    return {
        "transactions_scanned": len(txs),
        "candidates_count": len(candidates),
        "candidates_amount": round(sum(c["amount"] for c in candidates), 2),
        "base": base,
        "potential_if_all_confirmed": all_potential,
        "fast_path": fast_path,
        "groups": groups,
        "candidates": sorted(candidates, key=lambda x: (x["year"], x["date"], x["amount"])),
    }


def render_index_html() -> str:
    counter = METRIKA_COUNTER_ID if METRIKA_COUNTER_ID.isdigit() else "0"
    return INDEX_HTML.replace("__METRIKA_COUNTER_ID__", counter)

@app.get("/", response_class=HTMLResponse)
def index():
    return HTMLResponse(render_index_html())


@app.get("/terms", response_class=HTMLResponse)
def terms():
    return HTMLResponse(terms_html())


@app.get("/privacy", response_class=HTMLResponse)
def privacy():
    return HTMLResponse(privacy_html())


@app.get("/consent", response_class=HTMLResponse)
def consent():
    return HTMLResponse(consent_html())


@app.get("/api/legal-status")
def legal_status():
    return {
        "legal_ready": legal_ready(),
        "operator_address_configured": bool(OPERATOR_ADDRESS),
        "terms_version": TERMS_VERSION,
        "privacy_version": PRIVACY_VERSION,
        "consent_version": CONSENT_VERSION,
        "audit_path_configured": bool(CONSENT_AUDIT_PATH),
    }




def increment_public_found_refund(value: float) -> None:
    # Only the aggregate total is persisted; no merchant or transaction-level data.
    amount = max(0.0, float(value or 0))
    try:
        with _analytics_connect() as con:
            con.execute("""
                UPDATE public_stats
                SET total_found_refund = total_found_refund + ?,
                    analyses_count = analyses_count + 1,
                    updated_at = ?
                WHERE id = 1
            """, (
                round(amount, 2),
                datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
            ))
            con.commit()
    except Exception:
        pass

def get_public_found_refund() -> float:
    try:
        with _analytics_connect() as con:
            row = con.execute(
                "SELECT total_found_refund FROM public_stats WHERE id=1"
            ).fetchone()
            return round(float(row["total_found_refund"] or 0), 2) if row else 0.0
    except Exception:
        return 0.0

@app.get("/api/public-stats")
def public_stats():
    return {"found_refund": get_public_found_refund()}

class AnalyticsEventPayload(BaseModel):
    event: str
    session_id: str
    utm_source: str = ""
    utm_medium: str = ""
    utm_campaign: str = ""
    utm_content: str = ""
    utm_term: str = ""
    referrer_host: str = ""

@app.post("/api/analytics/event")
def analytics_event(payload: AnalyticsEventPayload):
    if payload.event not in ANALYTICS_EVENTS:
        raise HTTPException(400, "Unknown analytics event.")
    if not valid_session_id(payload.session_id):
        raise HTTPException(400, "Invalid session.")
    record_analytics_event(
        event=payload.event,
        session_id=payload.session_id,
        attribution=analytics_attribution_from_form(
            payload.utm_source, payload.utm_medium, payload.utm_campaign,
            payload.utm_content, payload.utm_term, payload.referrer_host,
        ),
    )
    return {"ok": True}

def require_analytics_token(request: Request):
    if not ANALYTICS_DASHBOARD_TOKEN:
        raise HTTPException(503, "Analytics dashboard token is not configured.")
    supplied = request.headers.get("x-analytics-token", "")
    if not hmac.compare_digest(supplied, ANALYTICS_DASHBOARD_TOKEN):
        raise HTTPException(401, "Unauthorized.")

@app.get("/api/analytics/summary")
def analytics_summary(request: Request, days: int = 30):
    require_analytics_token(request)
    days = max(1, min(int(days), 3650))
    start_ts = datetime.utcfromtimestamp(time.time() - days * 86400).replace(microsecond=0).isoformat() + "Z"
    with _analytics_connect() as con:
        event_rows = con.execute("""
            SELECT event, COUNT(DISTINCT session_id) AS sessions, COUNT(*) AS events
            FROM analytics_events
            WHERE created_at >= ?
            GROUP BY event
        """, (start_ts,)).fetchall()
        source_rows = con.execute("""
            SELECT
                CASE WHEN utm_source IS NULL OR utm_source='' THEN 'direct' ELSE utm_source END AS source,
                COUNT(DISTINCT CASE WHEN event='visit' THEN session_id END) AS visits,
                COUNT(DISTINCT CASE WHEN event='analysis_success' THEN session_id END) AS analyses,
                COUNT(DISTINCT CASE WHEN event='result_view' THEN session_id END) AS results,
                COUNT(DISTINCT CASE WHEN event='payment_success' THEN session_id END) AS paid_sessions,
                SUM(CASE WHEN event='payment_success' THEN COALESCE(payment_amount,0) ELSE 0 END) AS revenue
            FROM analytics_events
            WHERE created_at >= ?
            GROUP BY source
            ORDER BY revenue DESC, paid_sessions DESC, visits DESC
        """, (start_ts,)).fetchall()
        payments = con.execute("""
            SELECT COUNT(*) AS payments, COALESCE(SUM(payment_amount),0) AS revenue
            FROM analytics_events
            WHERE created_at >= ? AND event='payment_success'
        """, (start_ts,)).fetchone()
        errors = con.execute("""
            SELECT parser, bank, COUNT(*) AS errors
            FROM analytics_events
            WHERE created_at >= ? AND event='analysis_error'
            GROUP BY parser, bank
            ORDER BY errors DESC
            LIMIT 20
        """, (start_ts,)).fetchall()

    events = {r["event"]: {"sessions": r["sessions"], "events": r["events"]} for r in event_rows}
    return {
        "days": days,
        "from": start_ts,
        "events": events,
        "payments": {"count": payments["payments"], "revenue": payments["revenue"]},
        "sources": [dict(r) for r in source_rows],
        "analysis_errors": [dict(r) for r in errors],
    }

def analytics_dashboard_html() -> str:
    return r"""<!doctype html>
<html lang="ru"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Аналитика — СделатьВычет</title>
<style>
*{box-sizing:border-box}body{margin:0;background:#f6f6f3;color:#111;font-family:Inter,system-ui,-apple-system,sans-serif}
.wrap{max-width:1180px;margin:0 auto;padding:32px 20px 70px}.top{display:flex;justify-content:space-between;gap:20px;align-items:center;margin-bottom:26px}
h1{margin:0;font-size:34px;letter-spacing:-.05em}.muted{color:#777;font-size:12px}.auth{display:flex;gap:8px}
input,select,button{font:inherit;border:1px solid #ddd;border-radius:9px;padding:10px;background:#fff}button{background:#111;color:#fff;font-weight:750;cursor:pointer}
.cards{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}.card{background:#fff;border:1px solid #e3e3dd;border-radius:14px;padding:17px}
.card span{font-size:9px;color:#777;text-transform:uppercase;letter-spacing:.08em;font-weight:800}.card b{display:block;font-size:27px;margin-top:8px;letter-spacing:-.045em}
.section{background:#fff;border:1px solid #e3e3dd;border-radius:16px;margin-top:14px;padding:20px}h2{font-size:18px;margin:0 0 14px}
.funnel{display:grid;grid-template-columns:repeat(7,1fr);gap:8px}.step{background:#f7f7f2;border-radius:10px;padding:12px}.step b{display:block;font-size:19px}.step span{font-size:9px;color:#777}
table{width:100%;border-collapse:collapse;font-size:11px}th,td{text-align:left;padding:10px;border-bottom:1px solid #eee}th{color:#777;font-size:9px;text-transform:uppercase;letter-spacing:.06em}
.good{color:#287a48}.warn{color:#9a6b12}.error{color:#bd3a32}
@media(max-width:850px){.cards{grid-template-columns:1fr 1fr}.funnel{grid-template-columns:1fr 1fr}table{min-width:680px}.tableWrap{overflow:auto}.top{align-items:flex-start;flex-direction:column}}
</style></head><body><div class="wrap">
<div class="top"><div><h1>СделатьВычет · аналитика</h1><div class="muted">First-party воронка. Без merchant names и содержимого банковских операций.</div></div>
<div class="auth"><select id="days"><option value="7">7 дней</option><option value="30" selected>30 дней</option><option value="90">90 дней</option><option value="365">365 дней</option></select><input id="token" type="password" placeholder="Analytics token"><button onclick="load()">Открыть</button></div></div>
<div id="status" class="muted">Введите ANALYTICS_DASHBOARD_TOKEN.</div>
<div id="app" style="display:none">
<div class="cards">
<div class="card"><span>Визиты</span><b id="visits">—</b></div>
<div class="card"><span>Успешные анализы</span><b id="analyses">—</b></div>
<div class="card"><span>Оплаты</span><b id="payments">—</b></div>
<div class="card"><span>Выручка</span><b id="revenue">—</b></div>
</div>
<div class="section"><h2>Воронка</h2><div class="funnel" id="funnel"></div></div>
<div class="section"><h2>Конверсии</h2><div class="cards">
<div class="card"><span>Visit → Analysis</span><b id="crVA">—</b></div>
<div class="card"><span>Result → Pay</span><b id="crRP">—</b></div>
<div class="card"><span>Visit → Pay</span><b id="crVP">—</b></div>
<div class="card"><span>Средний платёж</span><b id="avgPay">—</b></div>
</div></div>
<div class="section"><h2>Источники</h2><div class="tableWrap"><table><thead><tr><th>Источник</th><th>Визиты</th><th>Анализы</th><th>Результаты</th><th>Оплаты</th><th>CR visit→pay</th><th>Выручка</th></tr></thead><tbody id="sources"></tbody></table></div></div>
</div></div>
<script>
const f=n=>new Intl.NumberFormat('ru-RU').format(n||0),rub=n=>f(n)+' ₽',pct=(a,b)=>b?((a/b)*100).toFixed(1)+'%':'0%';
function ev(d,n){return d.events?.[n]?.sessions||0}
async function load(){
 const token=document.getElementById('token').value.trim(),days=document.getElementById('days').value;
 sessionStorage.setItem('sv_analytics_token',token);
 const st=document.getElementById('status');st.textContent='Загружаю…';
 try{
  const r=await fetch('/api/analytics/summary?days='+days,{headers:{'X-Analytics-Token':token}});
  const d=await r.json();if(!r.ok)throw new Error(d.detail||'Ошибка');
  document.getElementById('app').style.display='block';st.textContent='Период: последние '+d.days+' дней';
  const visits=ev(d,'visit'),analyses=ev(d,'analysis_success'),results=ev(d,'result_view'),pays=ev(d,'payment_success');
  document.getElementById('visits').textContent=f(visits);document.getElementById('analyses').textContent=f(analyses);
  document.getElementById('payments').textContent=f(d.payments.count);document.getElementById('revenue').textContent=rub(d.payments.revenue);
  const steps=[['Визит','visit'],['Выбор файла','file_selected'],['Анализ','analysis_success'],['Результат','result_view'],['Нажал купить','payment_click'],['Оплатил','payment_success'],['Открыл отчёт','report_view']];
  document.getElementById('funnel').innerHTML=steps.map(x=>`<div class="step"><b>${f(ev(d,x[1]))}</b><span>${x[0]}</span></div>`).join('');
  document.getElementById('crVA').textContent=pct(analyses,visits);document.getElementById('crRP').textContent=pct(pays,results);document.getElementById('crVP').textContent=pct(pays,visits);
  document.getElementById('avgPay').textContent=rub(d.payments.count?d.payments.revenue/d.payments.count:0);
  document.getElementById('sources').innerHTML=(d.sources||[]).map(s=>`<tr><td><b>${s.source}</b></td><td>${f(s.visits)}</td><td>${f(s.analyses)}</td><td>${f(s.results)}</td><td>${f(s.paid_sessions)}</td><td>${pct(s.paid_sessions,s.visits)}</td><td>${rub(s.revenue)}</td></tr>`).join('');
 }catch(e){st.textContent=e.message;document.getElementById('app').style.display='none'}
}
document.getElementById('token').value=sessionStorage.getItem('sv_analytics_token')||'';
if(document.getElementById('token').value)load();
</script></body></html>"""

@app.get("/analytics", response_class=HTMLResponse)
def analytics_dashboard():
    return HTMLResponse(analytics_dashboard_html())

@app.get("/health")
def health():
    return {"ok": True, "version": "2.4.2", "legal_ready": legal_ready(), "service": "СделатьВычет", "metrika_configured": bool(METRIKA_COUNTER_ID), "analytics_db": bool(ANALYTICS_DB_PATH)}


@app.post("/api/analyze")
async def analyze(
    request: Request,
    file: UploadFile = File(...),
    terms_accepted: str = Form(""),
    pd_consent: str = Form(""),
    terms_version: str = Form(""),
    consent_version: str = Form(""),
    analytics_session_id: str = Form(""),
    utm_source: str = Form(""),
    utm_medium: str = Form(""),
    utm_campaign: str = Form(""),
    utm_content: str = Form(""),
    utm_term: str = Form(""),
    referrer_host: str = Form(""),
):
    accepted = str(terms_accepted).lower() in {"1", "true", "yes", "on"}
    consented = str(pd_consent).lower() in {"1", "true", "yes", "on"}
    analytics_session_id = clean_analytics_value(analytics_session_id, 80)
    attribution = analytics_attribution_from_form(
        utm_source, utm_medium, utm_campaign, utm_content, utm_term, referrer_host
    )

    if not accepted:
        raise HTTPException(400, "Перед загрузкой необходимо отдельно принять Пользовательское соглашение.")
    if not consented:
        raise HTTPException(400, "Перед загрузкой необходимо отдельно дать согласие на обработку персональных данных.")
    if terms_version != TERMS_VERSION or consent_version != CONSENT_VERSION:
        raise HTTPException(409, "Юридические документы обновились. Обновите страницу и подтвердите актуальную редакцию.")

    data = await file.read()
    if not data:
        raise HTTPException(400, "Файл пустой.")
    if len(data) > 30 * 1024 * 1024:
        raise HTTPException(413, "Файл больше 30 МБ. Загрузите более компактную выписку.")
    name = (file.filename or "").lower()
    bank_key = "unknown"
    parser = ""
    try:
        if name.endswith(".pdf"):
            pdf_text = extract_pdf_text(data)
            bank = detect_pdf_bank_from_text(pdf_text)
            bank_key, bank_display = recognize_bank(pdf_text)
            if bank == "sber":
                txs, parser = parse_sber_pdf_text(pdf_text), "Сбер PDF"
            elif bank == "tbank":
                txs, parser = parse_tbank_pdf_text(pdf_text), "Т-Банк PDF"
            elif bank == "alfa":
                txs, parser = parse_alfa_pdf_text(pdf_text), "Альфа-Банк PDF"
            elif bank == "vtb":
                try:
                    txs = parse_vtb_pdf_text(pdf_text)
                    parser = "ВТБ PDF"
                except ValueError:
                    txs = parse_generic_tax_pdf_text(pdf_text, bank_display)
                    parser = "ВТБ PDF · универсальный разбор"
            elif bank == "raiffeisen":
                try:
                    txs = parse_raiffeisen_pdf_text(pdf_text)
                    parser = "Райффайзен PDF"
                except ValueError:
                    txs = parse_generic_tax_pdf_text(pdf_text, bank_display)
                    parser = "Райффайзен PDF · универсальный разбор"
            elif bank.startswith("known:"):
                txs = parse_generic_tax_pdf_text(pdf_text, bank_display)
                parser = f"{bank_display} PDF · универсальный разбор"
            else:
                txs = parse_generic_tax_pdf_text(pdf_text, "Неизвестный банк")
                parser = "PDF · универсальный разбор"
        elif name.endswith(".csv"):
            txs, parser = parse_csv(data), "CSV"
        elif name.endswith(".xlsx") or name.endswith(".xlsm"):
            txs, parser = parse_xlsx(data), "XLSX"
        else:
            raise ValueError("Поддерживаются PDF, CSV и XLSX.")

        result = analyze_transactions(txs)
        result["filename"], result["parser"] = file.filename, parser

        analysis_id = uuid.uuid4().hex
        consent_event = {
            "event_id": uuid.uuid4().hex,
            "analysis_id": analysis_id,
            "created_at_utc": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
            "event": "statement_upload_consent",
            "terms_accepted": True,
            "pd_consent": True,
            "terms_version": TERMS_VERSION,
            "privacy_version": PRIVACY_VERSION,
            "consent_version": CONSENT_VERSION,
            "ip_hash": privacy_hash(request_ip(request)),
            "user_agent_hash": privacy_hash(request.headers.get("user-agent", "")),
            "operator_inn": OPERATOR_INN,
        }
        append_consent_audit(consent_event)

        cleanup_analyses()
        ANALYSES[analysis_id] = {
            "created_at": time.time(),
            "result": result,
            "paid": False,
            "payment_id": None,
            "analytics_session_id": analytics_session_id if valid_session_id(analytics_session_id) else "",
            "analytics_attribution": attribution,
            "analytics_payment_success_recorded": False,
            "analytics_report_view_recorded": False,
            "legal": {
                "event_id": consent_event["event_id"],
                "terms_version": TERMS_VERSION,
                "privacy_version": PRIVACY_VERSION,
                "consent_version": CONSENT_VERSION,
            },
        }

        if valid_session_id(analytics_session_id):
            record_analytics_event(
                event="analysis_success",
                session_id=analytics_session_id,
                attribution=attribution,
                bank=bank_key,
                parser=parser,
                candidate_count=result["candidates_count"],
                refund_value=result["potential_if_all_confirmed"]["refund_from"],
                dedupe_key=f"analysis_success:{analysis_id}",
            )

        increment_public_found_refund(result["potential_if_all_confirmed"]["refund_from"])

        # ВАЖНО: до оплаты браузер получает только агрегаты, без категорий и транзакций.
        return JSONResponse({
            "analysis_id": analysis_id,
            "expenses_found": result["candidates_amount"],
            "refund_from": result["potential_if_all_confirmed"]["refund_from"],
            "price": REPORT_PRICE_RUB,
        })
    except ValueError as e:
        if valid_session_id(analytics_session_id):
            record_analytics_event(
                event="analysis_error", session_id=analytics_session_id,
                attribution=attribution, bank=bank_key, parser=parser or "parse_error"
            )
        raise HTTPException(422, str(e))
    except HTTPException:
        raise
    except Exception as e:
        if valid_session_id(analytics_session_id):
            record_analytics_event(
                event="analysis_error", session_id=analytics_session_id,
                attribution=attribution, bank=bank_key, parser=parser or "server_error"
            )
        raise HTTPException(500, f"Ошибка обработки: {e}")


class PaymentPayload(BaseModel):
    analysis_id: str


@app.post("/api/create-payment")
def create_payment(payload: PaymentPayload, request: Request):
    item = get_analysis_or_404(payload.analysis_id)
    sid = item.get("analytics_session_id", "")
    attr = item.get("analytics_attribution", {})
    if valid_session_id(sid):
        record_analytics_event(
            event="payment_created", session_id=sid, attribution=attr,
            dedupe_key=f"payment_created:{payload.analysis_id}"
        )
    if item.get("paid"):
        return {"status": "succeeded", "confirmation_url": f"{public_base_url(request)}/?analysis={payload.analysis_id}&paid=1"}

    if PAYMENT_TEST_MODE:
        payment_id = "test_" + uuid.uuid4().hex
        item["payment_id"] = payment_id
        return {
            "status": "pending",
            "test_mode": True,
            "confirmation_url": f"{public_base_url(request)}/api/test-pay?analysis_id={payload.analysis_id}&payment_id={payment_id}",
        }

    return_url = f"{public_base_url(request)}/?analysis={payload.analysis_id}&payment=return"
    payment = yookassa_request(
        "POST",
        "/payments",
        {
            "amount": {"value": f"{REPORT_PRICE_RUB:.2f}", "currency": "RUB"},
            "capture": True,
            "confirmation": {"type": "redirect", "return_url": return_url},
            "description": "СделатьВычет — персональный налоговый отчёт",
            "metadata": {"analysis_id": payload.analysis_id},
        },
        idempotence_key=str(uuid.uuid4()),
    )
    item["payment_id"] = payment.get("id")
    confirmation_url = (payment.get("confirmation") or {}).get("confirmation_url")
    if not confirmation_url:
        raise HTTPException(502, "ЮKassa не вернула ссылку на оплату.")
    return {"status": payment.get("status", "pending"), "confirmation_url": confirmation_url}


@app.get("/api/test-pay")
def test_pay(analysis_id: str, payment_id: str):
    if not PAYMENT_TEST_MODE:
        raise HTTPException(404, "Тестовая оплата отключена.")
    item = get_analysis_or_404(analysis_id)
    if item.get("payment_id") != payment_id:
        raise HTTPException(400, "Некорректный тестовый платёж.")
    item["paid"] = True
    sid = item.get("analytics_session_id", "")
    if valid_session_id(sid) and not item.get("analytics_payment_success_recorded"):
        record_analytics_event(
            event="payment_success", session_id=sid,
            attribution=item.get("analytics_attribution", {}),
            payment_amount=REPORT_PRICE_RUB,
            dedupe_key=f"payment_success:{analysis_id}",
        )
        item["analytics_payment_success_recorded"] = True
    return RedirectResponse(url=f"/?analysis={analysis_id}&paid=1", status_code=302)


@app.get("/api/payment-status")
def payment_status(analysis_id: str):
    item = get_analysis_or_404(analysis_id)
    if item.get("paid"):
        return {"status": "succeeded", "paid": True}

    payment_id = item.get("payment_id")
    if not payment_id:
        return {"status": "not_created", "paid": False}

    if PAYMENT_TEST_MODE and str(payment_id).startswith("test_"):
        return {"status": "pending", "paid": False}

    payment = yookassa_request("GET", f"/payments/{payment_id}")
    succeeded = payment.get("status") == "succeeded" and bool(payment.get("paid"))
    if succeeded:
        item["paid"] = True
        sid = item.get("analytics_session_id", "")
        if valid_session_id(sid) and not item.get("analytics_payment_success_recorded"):
            record_analytics_event(
                event="payment_success", session_id=sid,
                attribution=item.get("analytics_attribution", {}),
                payment_amount=REPORT_PRICE_RUB,
                dedupe_key=f"payment_success:{analysis_id}",
            )
            item["analytics_payment_success_recorded"] = True
    return {"status": payment.get("status"), "paid": succeeded}


@app.get("/api/report/{analysis_id}")
def paid_report(analysis_id: str):
    item = get_analysis_or_404(analysis_id)
    if not item.get("paid"):
        raise HTTPException(402, "Сначала оплатите полный отчёт.")
    sid = item.get("analytics_session_id", "")
    if valid_session_id(sid) and not item.get("analytics_report_view_recorded"):
        record_analytics_event(
            event="report_view", session_id=sid,
            attribution=item.get("analytics_attribution", {}),
            dedupe_key=f"report_view:{analysis_id}",
        )
        item["analytics_report_view_recorded"] = True
    return JSONResponse(item["result"])


class RecalcPayload(BaseModel):
    candidates: list[dict[str, Any]]
    selected_ids: list[str]


@app.post("/api/recalculate")
def recalculate(payload: RecalcPayload):
    return calculate_refund(payload.candidates, set(payload.selected_ids))


class PacketPayload(BaseModel):
    filename: str = "Выписка"
    candidates: list[dict[str, Any]]
    selected_ids: list[str]


@app.post("/api/packet")
def packet(payload: PacketPayload):
    chosen = set(payload.selected_ids)
    selected = [c for c in payload.candidates if c["id"] in chosen]
    calc = calculate_refund(payload.candidates, chosen)

    easy_cats = {"medicine", "fitness", "education"}
    question_cats = {"insurance"}
    extra_cats = {"pharmacy", "donation"}

    easy = [c for c in selected if c["category"] in easy_cats]
    questions = [c for c in selected if c["category"] in question_cats]
    extras = [c for c in selected if c["category"] in extra_cats]

    def calc_subset(items):
        return calculate_refund(items, {c["id"] for c in items}) if items else {"refund_from": 0, "selected_amount": 0}

    easy_calc = calc_subset(easy)
    question_calc = calc_subset(questions)

    def unique_merchants(items):
        out = []
        for c in items:
            m = (c.get("merchant") or "").strip()
            if m and m != "Не удалось определить" and m not in out:
                out.append(m)
        return out

    def short_list(items, limit=8):
        names = unique_merchants(items)
        shown = names[:limit]
        tail = f" + ещё {len(names)-limit}" if len(names) > limit else ""
        return ", ".join(escape(x) for x in shown) + tail if names else "организации из найденных операций"

    rows = "".join(
        f"<tr><td>{escape(c['date'])}</td><td>{escape(c['category_name'])}</td>"
        f"<td>{escape(c['merchant'])}</td><td>{c['amount']:,.2f} ₽</td></tr>"
        for c in selected
    )

    fast_ids = {c['id'] for c in easy + questions}
    fast_calc = calculate_refund(payload.candidates, fast_ids) if fast_ids else {"refund_from": 0}
    extra_increment = max(0, calc['refund_from'] - fast_calc['refund_from'])
    action_count = int(bool(easy)) + int(bool(questions))

    action_html = []
    if easy:
        action_html.append(
            f"<div class='action'><div class='n'>1</div><div><b>Получить справки одним пакетом — от {easy_calc['refund_from']:,.0f} ₽</b>"
            f"<p>Обратиться в {len(unique_merchants(easy))} организаций: {short_list(easy)}. "
            f"СделатьВычет уже собрал оплаты и разбил их по организациям.</p></div></div>"
        )
    if questions:
        n = 2 if easy else 1
        action_html.append(
            f"<div class='action'><div class='n'>{n}</div><div><b>Ответить на {len(unique_merchants(questions))} коротких вопроса — до {question_calc['refund_from']:,.0f} ₽</b>"
            f"<p>Уточнить тип договора: {short_list(questions)}. Если это неподходящий тип страховки — просто исключить оплату.</p></div></div>"
        )

    extra_html = ""
    if extras:
        extra_html = f"<div class='extra'><b>Можно попробовать вернуть ещё до {extra_increment:,.0f} ₽</b><p>Лекарства и другие расходы требуют больше документов. Их можно оставить на потом.</p></div>"

    html = f'''<!doctype html>
<html lang="ru"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>СделатьВычет — персональный пакет</title>
<style>
:root{{--text:#111827;--muted:#667085;--line:#e5e9f0;--green:#07864c;--greenSoft:#eaf8f1;--brand:#3157e7;--bg:#f5f7fb}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--bg);color:var(--text);font-family:Inter,Arial,sans-serif;line-height:1.45}}.page{{max-width:920px;margin:36px auto;padding:0 18px 60px}}.top{{display:flex;align-items:center;justify-content:space-between;margin-bottom:18px}}.brand{{font-size:20px;font-weight:900}}.brand i{{display:inline-grid;place-items:center;width:34px;height:34px;border-radius:11px;background:linear-gradient(145deg,#3157e7,#6c4df6);color:white;font-style:normal;margin-right:8px}}.source{{font-size:11px;color:var(--muted)}}.hero{{background:linear-gradient(135deg,#121a2f,#22265a);color:#fff;border-radius:24px;padding:28px;box-shadow:0 18px 50px rgba(28,39,60,.12)}}.hero small{{color:#bbc5d9}}.big{{font-size:44px;line-height:1;font-weight:950;letter-spacing:-.045em;margin:6px 0 10px}}.fast{{color:#b8f0d0;font-weight:800;font-size:13px}}h2{{font-size:21px;letter-spacing:-.025em;margin:26px 0 12px}}.action{{display:grid;grid-template-columns:40px 1fr;gap:12px;border:1px solid var(--line);border-radius:16px;padding:15px;margin:9px 0;background:#fff}}.n{{width:36px;height:36px;border-radius:11px;background:#eef2ff;color:#4452ce;display:grid;place-items:center;font-weight:900}}.action b{{font-size:14px}}.action p,.extra p{{color:var(--muted);font-size:12px;margin:4px 0 0}}.extra{{background:#fffaf0;border:1px solid #f0e0bb;border-radius:16px;padding:15px;margin:14px 0}}.extra b{{color:#896100}}.ops{{background:#fff;border:1px solid var(--line);border-radius:16px;overflow:hidden}}table{{width:100%;border-collapse:collapse}}td,th{{padding:10px 9px;border-bottom:1px solid #edf0f4;text-align:left;font-size:11px}}th{{font-size:9px;text-transform:uppercase;letter-spacing:.06em;color:#7c8799}}.note{{background:#f8fafc;border:1px solid var(--line);padding:13px 14px;border-radius:13px;color:#6b7280;font-size:10px;line-height:1.5;margin-top:18px}}@media(max-width:600px){{.page{{margin-top:16px}}.big{{font-size:36px}}.hero{{padding:22px}}th:nth-child(1),td:nth-child(1){{display:none}}}}
</style></head><body><div class="page">
<div class="top"><div class="brand">
  <svg class="brandMark" viewBox="0 0 44 44" fill="none" xmlns="http://www.w3.org/2000/svg" aria-label="СделатьВычет">
    <rect x="1" y="1" width="42" height="42" rx="13" fill="#B7FF2A"/>
    <path d="M14 23.5L19.2 28.7L30 16.8" stroke="#142000" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/>
  </svg>
  <div class="brandWord"><b>СделатьВычет</b><span>Налоговый помощник</span></div>
</div><div class="source">Источник: {escape(payload.filename)}</div></div>
<div class="hero"><small>Ваш потенциальный возврат</small><div class="big">от {calc['refund_from']:,.0f} ₽</div><div class="fast">До {fast_calc['refund_from']:,.0f} ₽ — за {action_count or 1} простых действия</div></div>
<h2>Быстрый путь</h2>{''.join(action_html) or '<p>Сначала выберите подходящие операции.</p>'}{extra_html}
<h2>Все включённые операции</h2><div class="ops"><table><thead><tr><th>Дата</th><th>Категория</th><th>Организация</th><th>Сумма</th></tr></thead><tbody>{rows}</tbody></table></div>
<div class="note"><b>Важно:</b> расчёт использует базовые 13% и консервативный общий социальный лимит 150 000 ₽ на год. Это персональный навигатор и рабочий пакет, а не гарантия возврата.</div>
</div></body></html>'''
    return Response(html, media_type="text/html; charset=utf-8", headers={"Content-Disposition": 'attachment; filename="tax_radar_packet.html"'})


INDEX_HTML = r'''<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<meta name="color-scheme" content="light"/>
<title>СделатьВычет — найдём ваш налоговый вычет</title>
<style>
:root{
  --bg:#f6f6f3;
  --surface:#ffffff;
  --text:#11110f;
  --muted:#74746d;
  --line:#e5e5df;
  --line2:#d8d8d1;
  --accent:#b7ff2a;
  --accentInk:#182000;
  --soft:#f0f0eb;
  --green:#287a48;
  --greenSoft:#edf7ef;
  --amber:#806615;
  --amberSoft:#f7f3df;
  --red:#bd3a32;
}
*{box-sizing:border-box}
html{scroll-behavior:smooth}
body{
  margin:0;background:var(--bg);color:var(--text);
  font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
  -webkit-font-smoothing:antialiased;
}
button,input,select{font:inherit}
button{touch-action:manipulation}
.wrap{max-width:1080px;margin:0 auto;padding:24px 20px 72px}
.header{display:flex;align-items:center;justify-content:space-between;padding:0 0 20px;border-bottom:1px solid var(--line)}
.brand{display:flex;align-items:center;gap:9px;font-weight:820;letter-spacing:-.025em;font-size:18px}
.brandMark{width:30px;height:30px;border-radius:8px;display:grid;place-items:center;background:#111;color:#fff;font-size:14px;font-weight:850}
.beta{font-size:9px;font-weight:800;color:var(--muted);border:1px solid var(--line);background:#fff;padding:4px 7px;border-radius:999px;text-transform:uppercase;letter-spacing:.07em}
.secure{display:flex;align-items:center;gap:7px;color:var(--muted);font-size:11px}
.secureDot{width:7px;height:7px;border-radius:50%;background:#29bd69}

.hero{
  display:grid;grid-template-columns:1.1fr .9fr;gap:56px;align-items:end;
  padding:68px 0 48px;border-bottom:1px solid var(--line);
}
.heroCopy,.preview{position:relative}
.eyebrow{display:flex;align-items:center;gap:8px;color:var(--muted);font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.11em;margin-bottom:17px}
.eyebrowDot{width:6px;height:6px;border-radius:50%;background:var(--accent);box-shadow:0 0 0 4px rgba(183,255,42,.15)}
h1{font-size:64px;line-height:.96;letter-spacing:-.06em;margin:0;max-width:720px;font-weight:820}
.heroText{font-size:17px;line-height:1.55;color:var(--muted);max-width:620px;margin:22px 0 27px}

.publicCounter{
  display:inline-flex;align-items:baseline;gap:7px;margin-top:18px;
  padding:10px 13px;border-radius:999px;background:#111;color:#fff;
  width:max-content;max-width:100%;box-shadow:0 8px 24px rgba(0,0,0,.08)
}
.publicCounter span{font-size:10px;color:#a8a8a2;font-weight:650}
.publicCounter b{font-size:14px;letter-spacing:-.03em;white-space:nowrap}
@media(max-width:600px){
  .publicCounter{padding:9px 11px;margin-top:15px}
  .publicCounter span{font-size:9px}
  .publicCounter b{font-size:12px}
}
.heroActions{display:flex;gap:13px;align-items:center;flex-wrap:wrap}
.btn{
  appearance:none;border:0;border-radius:10px;padding:12px 15px;font-weight:780;cursor:pointer;
  transition:transform .16s ease,opacity .16s ease;display:inline-flex;align-items:center;justify-content:center;gap:7px;text-decoration:none;
  background:#111;color:#fff;
}
.btn:hover{transform:translateY(-1px)}
.btnPrimary{background:var(--accent);color:var(--accentInk)}
.btnBrand{background:#111;color:#fff}
.btnGhost{background:#fff;color:#222;border:1px solid var(--line)}
.btn:disabled{opacity:.42;cursor:not-allowed;transform:none}
.heroHint{font-size:11px;color:var(--muted)}

.preview{background:#111;color:#fff;border-radius:18px;padding:23px;box-shadow:0 18px 42px rgba(0,0,0,.08)}
.previewTop{display:flex;justify-content:space-between;align-items:center;margin-bottom:28px;color:#989894;font-size:10px;text-transform:uppercase;letter-spacing:.07em}
.previewPill{border:1px solid #2d2d2b;padding:5px 7px;border-radius:999px;color:#b8b8b4}
.previewLabel{font-size:11px;color:#9c9c97}
.previewMoney{font-size:47px;font-weight:820;letter-spacing:-.05em;margin:5px 0 9px}
.previewFast{display:flex;align-items:center;gap:8px;color:#deded8;font-size:12px;font-weight:700;margin-bottom:22px}
.previewFast i{width:20px;height:20px;display:grid;place-items:center;border-radius:50%;font-style:normal;background:var(--accent);color:#151a00;font-size:11px}
.previewSteps{display:grid;grid-template-columns:1fr 1fr;gap:8px}
.previewStep{background:#191917;border:1px solid #282825;border-radius:11px;padding:12px;font-size:11px;color:#bcbcb6;line-height:1.4}
.previewStep b{display:block;color:#fff;font-size:12px;margin-bottom:4px}

.uploaderCard{margin-top:26px;background:#fff;border:1px solid var(--line);border-radius:18px;padding:26px}
.uploaderHead{display:flex;justify-content:space-between;gap:20px;align-items:flex-start;margin-bottom:15px}
.uploaderHead h2{font-size:21px;letter-spacing:-.03em;margin:0 0 5px}
.uploaderHead p{margin:0;color:var(--muted);font-size:11px;line-height:1.5;max-width:650px}
.formatPills{display:flex;gap:5px;flex-wrap:wrap;justify-content:flex-end}
.formatPill{border:1px solid var(--line);background:#fafaf7;border-radius:999px;padding:5px 7px;color:var(--muted);font-size:9px;font-weight:800}
.upload{
  border:1px solid var(--line);border-radius:14px;background:#fafaf7;padding:18px;
  cursor:pointer;transition:.16s ease;display:flex;align-items:center;justify-content:space-between;gap:16px;text-align:left
}
.upload:hover,.upload.drag{border-color:#b9b9b0;background:#f7f7f2}
.upload.selected{border-color:#b9b9b0;background:#fff}
.uploadLeft{display:flex;align-items:center;gap:13px;min-width:0}
.uploadIcon{width:42px;height:42px;flex:0 0 42px;border:1px solid var(--line);background:#fff;border-radius:11px;display:grid;place-items:center;font-size:18px}
.uploadMeta{min-width:0}
.uploadTitle{font-weight:780;font-size:13px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:560px}
.uploadSub{color:var(--muted);font-size:10px;margin-top:3px}
.uploadPick{flex:0 0 auto;background:#111;color:#fff;border-radius:9px;padding:9px 11px;font-size:10px;font-weight:800}
.upload.selected .uploadPick{background:#f1f1ed;color:#333}

input[type=file]{display:none}

.legalConsent{
  margin-top:14px;padding:14px 15px;border:1px solid var(--line);
  background:#fafaf7;border-radius:11px;display:grid;gap:10px
}
.legalRow{display:flex;gap:9px;align-items:flex-start;font-size:9px;line-height:1.45;color:#555}
.legalRow input{margin:2px 0 0;width:15px;height:15px;accent-color:#111;flex:0 0 auto}
.legalRow a{color:#111;text-decoration:underline;text-underline-offset:2px}
.legalMini{font-size:8px;color:#96968f;line-height:1.5;margin-top:1px}
.pill{display:inline-flex;align-items:center;padding:7px 10px;border:1px solid var(--line);border-radius:999px;background:#fff;font-size:8px;color:#666;font-weight:800;text-transform:uppercase;letter-spacing:.06em}

/* 2.3 polish */
.brand{display:flex;align-items:center;gap:10px}
.brandLogo{width:34px;height:34px;display:block;flex:0 0 auto}
.brandName{font-weight:880;letter-spacing:-.045em;font-size:18px}
.beta{padding:4px 6px;border-radius:999px;background:#ededE7;color:#7a7a73;font-size:7px;letter-spacing:.07em;text-transform:uppercase;font-weight:850}
.hero{padding-top:26px}
.hero h1{max-width:760px;text-wrap:balance}
.heroText{max-width:650px}
.preview{box-shadow:0 18px 60px rgba(0,0,0,.08)}
.uploaderCard,.summaryCard,.resultShell{box-shadow:0 12px 45px rgba(0,0,0,.035)}
.upload{transition:border-color .18s ease,background .18s ease,transform .18s ease}
.upload:hover{border-color:#bbbcb4;background:#fdfdf9}
.upload.selected{border-color:#a7c967;background:#fbfff4}
.btnPrimary,.btnBrand,.payBtn{transition:transform .15s ease,opacity .15s ease,box-shadow .15s ease}
.btnPrimary:hover,.btnBrand:hover,.payBtn:hover{transform:translateY(-1px);box-shadow:0 8px 24px rgba(0,0,0,.10)}
.summaryNumber strong{letter-spacing:-.045em}
.resultTop{background:linear-gradient(180deg,#fff 0%,#fdfdf9 100%)}
.fastBox{border:1px solid #dfecc2;background:#f8ffe9}
.metricTiles{grid-template-columns:repeat(3,minmax(0,1fr))}
.metricTile{min-width:0}
.metricTile .value{white-space:nowrap}
.journey{box-shadow:0 10px 35px rgba(0,0,0,.035)}
.guideStep{transition:background .15s ease}
.guideStep:hover{background:#fdfdf9}
.detailsCard{overflow:hidden}
.legalConsent{background:#fbfbf7}

.mobileSticky{display:none}

@media(max-width:900px){
  body{overflow-x:hidden}
  .wrap{padding-left:16px!important;padding-right:16px!important}
  header{padding-top:16px}
  .secure{font-size:8px}
  .brandName{font-size:16px}
  .brandLogo{width:31px;height:31px}
  .hero{grid-template-columns:1fr!important;gap:20px!important;padding-top:22px}
  .hero h1{font-size:44px!important;line-height:.96!important}
  .heroText{font-size:12px!important;line-height:1.6!important}
  .heroActions{align-items:flex-start!important;flex-direction:column!important}
  .preview{min-height:0!important}
  .proofStrip,
  
  .uploaderHead{display:flex!important;flex-direction:column!important;align-items:flex-start!important;gap:12px!important}
  .formatPills{margin-left:0!important}
  .upload{padding:15px!important}
  .uploadLeft{min-width:0}
  .uploadMeta{min-width:0}
  .uploadTitle{white-space:normal!important}
  .uploadSub{white-space:normal!important;line-height:1.4}
  .uploadPick{flex:0 0 auto}
  .uploadBottom{flex-direction:column!important;align-items:stretch!important}
  .uploadBottom .btn{width:100%}
  .parser{text-align:center}
  .summaryTop{padding:22px!important}
  .summaryNumbers{grid-template-columns:1fr!important}
  .summaryNumber+.summaryNumber{border-left:0!important;border-top:1px solid var(--line)}
  .paywall{grid-template-columns:1fr!important;gap:18px!important}
  .paywallPrice{justify-content:space-between}
  .payBtn{min-height:46px}
  .resultTop{padding:22px!important}
  .resultGrid{grid-template-columns:1fr!important;gap:18px!important}
  .refund{font-size:52px!important}
  .metricTiles{grid-template-columns:1fr!important}
  .resultBody{padding-left:18px!important;padding-right:18px!important}
  .journeyHero{grid-template-columns:1fr!important}
  .journeyMoney{text-align:left!important}
  .sectionTitleRow{align-items:flex-start!important;flex-direction:column!important}
  .actionCard{grid-template-columns:34px 1fr!important}
  .actionMoney{grid-column:2;text-align:left!important}
  .packetRow{align-items:flex-start!important;flex-direction:column!important}
  .packetRow .btn{width:100%}
  table{font-size:9px}
  .detailsInner{overflow-x:auto}
  .legalFooter{padding-bottom:84px!important}
  .mobileSticky{
    position:fixed;display:flex;left:12px;right:12px;bottom:12px;z-index:30;
    align-items:center;justify-content:space-between;gap:12px;padding:11px 12px;
    background:rgba(17,17,17,.94);backdrop-filter:blur(14px);
    border:1px solid #333;border-radius:15px;box-shadow:0 12px 35px rgba(0,0,0,.25);
    color:#fff
  }
  .mobileSticky span{font-size:9px;color:#aaa;display:block}
  .mobileSticky b{font-size:12px}
  .mobileSticky button{border:0;border-radius:9px;background:var(--accent);color:#172000;padding:10px 12px;font-size:9px;font-weight:850}
}
@media(max-width:520px){
  .wrap{padding-top:0!important}
  header{gap:8px}
  .secure{max-width:130px;text-align:right;line-height:1.3}
  .hero h1{font-size:38px!important}
  .proofCard{padding:14px}
  
  .legalRow{font-size:9px}
}

.proofStrip{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:18px 0 20px}
.proofCard{display:flex;gap:12px;align-items:flex-start;background:#fff;border:1px solid var(--line);border-radius:16px;padding:16px}
.proofNum{width:28px;height:28px;border-radius:50%;display:grid;place-items:center;background:#111;color:#fff;font-size:12px;font-weight:800;flex:0 0 auto}
.proofCard b{display:block;font-size:12px;letter-spacing:-.02em}
.proofCard span{display:block;margin-top:4px;font-size:10px;color:var(--muted);line-height:1.5}










.successBanner{display:none;margin:0 0 14px;padding:16px 18px;border-radius:14px;background:#efffcb;color:#334a00;border:1px solid #d9f0a2}
.successBanner.show{display:block}
.successBanner b{display:block;font-size:13px;letter-spacing:-.02em}
.successBanner p{margin:5px 0 0;font-size:10px;line-height:1.55}
.metricTiles{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:18px}
.metricTile{padding:16px;border:1px solid var(--line);border-radius:13px;background:#fafaf7}
.metricTile .label{font-size:9px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;font-weight:800}
.metricTile .value{margin-top:7px;font-size:24px;letter-spacing:-.04em;font-weight:820}
.metricTile .meta{margin-top:4px;color:var(--muted);font-size:9px;line-height:1.5}
@media(max-width:900px){.proofStrip,.metricTiles{grid-template-columns:1fr}}

.legalFooter{margin-top:34px;padding:20px 0;border-top:1px solid var(--line);display:flex;gap:16px;flex-wrap:wrap;font-size:9px;color:#777}
.legalFooter a{color:#555;text-decoration:none}.legalFooter a:hover{text-decoration:underline}

.uploadBottom{display:flex;align-items:center;gap:11px;margin-top:13px}
.parser{font-size:10px;color:var(--muted)}
.loader{display:none;margin-top:14px}
.loaderTop{display:flex;justify-content:space-between;gap:10px;align-items:center;color:var(--muted);font-size:10px;margin-bottom:7px}
.loaderTop b{color:#222;font-size:11px}
.track{height:3px;background:#ecece7;border-radius:99px;overflow:hidden}
.bar{height:100%;width:28%;background:#111;border-radius:99px;animation:scan 1.05s infinite alternate ease-in-out}
@keyframes scan{from{transform:translateX(-110%)}to{transform:translateX(360%)}}

.analyticsConsent{
  position:fixed;left:18px;bottom:18px;z-index:60;max-width:420px;
  padding:14px;background:#111;color:#fff;border:1px solid #333;border-radius:14px;
  box-shadow:0 18px 50px rgba(0,0,0,.28);display:none
}
.analyticsConsent.show{display:block}
.analyticsConsent b{font-size:11px}
.analyticsConsent p{margin:5px 0 11px;color:#aaa;font-size:9px;line-height:1.5}
.analyticsConsentActions{display:flex;gap:7px}
.analyticsConsent button{border:0;border-radius:8px;padding:8px 10px;font-size:9px;font-weight:800;cursor:pointer}
.analyticsConsent .allow{background:var(--accent);color:#172000}
.analyticsConsent .deny{background:#292927;color:#ddd}
@media(max-width:600px){.analyticsConsent{left:12px;right:12px;bottom:78px;max-width:none}}

#error{display:none;margin-top:12px;background:#fff1ef;border:1px solid #f0d5d1;color:var(--red);padding:10px 11px;border-radius:9px;font-size:11px}

.summary{display:none;margin-top:24px}
.summaryCard{background:#fff;border:1px solid var(--line);border-radius:18px;overflow:hidden}
.summaryTop{padding:30px;border-bottom:1px solid var(--line)}
.summaryBadge{display:inline-flex;align-items:center;gap:7px;color:var(--muted);font-size:9px;font-weight:800;text-transform:uppercase;letter-spacing:.1em}
.summaryBadge i{width:18px;height:18px;border-radius:50%;display:grid;place-items:center;background:var(--accent);color:#172000;font-style:normal}
.summaryTitle{font-size:22px;letter-spacing:-.035em;margin:18px 0 4px}
.summarySub{font-size:11px;color:var(--muted)}
.summaryNumbers{display:grid;grid-template-columns:1fr 1fr;gap:0;margin-top:26px;border:1px solid var(--line);border-radius:13px;overflow:hidden}
.summaryNumber{padding:20px}
.summaryNumber+ .summaryNumber{border-left:1px solid var(--line)}
.summaryNumber span{display:block;color:var(--muted);font-size:10px;margin-bottom:5px}
.summaryNumber b{font-size:34px;letter-spacing:-.045em}
.paywall{padding:24px 30px;display:grid;grid-template-columns:1fr auto;gap:24px;align-items:center;background:#111;color:#fff}
.paywallLabel{color:#a6a6a0;font-size:9px;text-transform:uppercase;letter-spacing:.09em;font-weight:800}
.paywallTitle{font-size:20px;font-weight:800;letter-spacing:-.03em;margin-top:5px}
.paywallText{color:#a9a9a3;font-size:10px;line-height:1.5;margin-top:6px;max-width:630px}
.paywallPrice{display:flex;align-items:center;gap:13px}
.price{font-size:28px;font-weight:820;letter-spacing:-.04em;white-space:nowrap}
.payBtn{background:var(--accent);color:var(--accentInk);border:0;border-radius:10px;padding:12px 16px;font-weight:820;cursor:pointer;white-space:nowrap}
.paySecure{font-size:9px;color:#8e8e89;margin-top:7px;text-align:right}
.paymentStatus{display:none;margin-top:12px;padding:10px 11px;border-radius:9px;background:#f6f6f1;color:var(--muted);font-size:10px}
.unlockBadge{display:inline-flex;align-items:center;gap:7px;color:var(--green);font-size:10px;font-weight:800;margin-bottom:12px}
#results{display:none;margin-top:24px}
.resultShell{background:#fff;border:1px solid var(--line);border-radius:16px;overflow:hidden}
.resultTop{padding:30px;border-bottom:1px solid var(--line)}
.resultBadge{display:flex;align-items:center;gap:6px;color:var(--muted);font-size:9px;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:19px}
.resultBadge i{width:16px;height:16px;border-radius:50%;display:grid;place-items:center;background:var(--accent);color:#172000;font-style:normal;font-size:9px}
.resultGrid{display:grid;grid-template-columns:1.18fr .82fr;gap:28px;align-items:end}
.resultLabel{font-size:11px;color:var(--muted);margin-bottom:5px}
.refund{font-size:68px;line-height:.95;letter-spacing:-.06em;font-weight:830}
.refund .from{font-size:20px;color:var(--muted);font-weight:650;letter-spacing:-.02em;margin-right:8px}
.resultMeta{margin-top:10px;color:var(--muted);font-size:10px;line-height:1.45}
.fastBox{border-left:1px solid var(--line);padding-left:26px}
.fastLabel{font-size:9px;color:var(--muted);font-weight:800;text-transform:uppercase;letter-spacing:.09em}
.fastMoney{font-size:31px;font-weight:820;letter-spacing:-.04em;margin-top:5px}
.fastText{color:var(--muted);font-size:10px;line-height:1.45;margin-top:6px}

.resultBody{padding:28px 30px 30px}
.sectionTitleRow{display:flex;justify-content:space-between;gap:18px;align-items:end}
.sectionTitleRow h2{font-size:21px;letter-spacing:-.03em;margin:0 0 4px}
.sectionTitleRow p{margin:0;color:var(--muted);font-size:11px}
.tiny{font-size:9px;color:var(--muted)}
.groups{display:flex;gap:6px;flex-wrap:wrap;margin:16px 0 5px}
.chip{border:1px solid var(--line);background:#fafaf7;border-radius:999px;padding:6px 8px;color:#55554f;font-size:9px;font-weight:700}

.actions{margin-top:14px}
.actionCard{display:grid;grid-template-columns:34px 1fr auto;gap:14px;align-items:start;padding:19px 0;border-top:1px solid var(--line)}
.actionNo{width:32px;height:32px;border-radius:9px;display:grid;place-items:center;background:var(--soft);color:#333;font-size:11px;font-weight:800}
.actionName{font-size:14px;font-weight:790;letter-spacing:-.02em}
.actionDesc{color:var(--muted);font-size:10px;line-height:1.5;margin-top:5px;max-width:660px}
.actionMoney{font-size:12px;font-weight:800;white-space:nowrap}
.actionBtn{margin-top:9px;border:0;background:transparent;padding:0;color:#111;font-size:10px;font-weight:800;cursor:pointer;text-decoration:underline;text-underline-offset:3px}
.batch,.questionBox{display:none;margin-top:12px;border-top:1px solid var(--line)}
.batchItem{padding:12px 0;border-bottom:1px solid var(--line);font-size:10px;color:var(--muted);line-height:1.45}
.batchItem b{display:block;color:#222;font-size:11px;margin-bottom:3px}
.copy{margin-top:6px;color:#111;font-weight:750;cursor:pointer}
.qrow{padding:11px 0;border-bottom:1px solid var(--line);display:grid;grid-template-columns:1fr minmax(220px,320px);gap:12px;align-items:center}
.qrow b{font-size:11px}
.qrow select{width:100%;border:1px solid var(--line);border-radius:8px;background:#fff;padding:8px 9px;color:#333;font-size:10px}

.extra{display:none;margin-top:18px;background:#111;color:#fff;border-radius:13px;padding:17px}
.extraHead{display:flex;justify-content:space-between;gap:16px;align-items:center}
.extraMoney{font-size:22px;font-weight:810;letter-spacing:-.035em}
.extraText{color:#aaa;font-size:10px;line-height:1.45;margin-top:4px}
.extraBtn{border:1px solid #353532;background:#191917;color:#eee;border-radius:8px;padding:8px 10px;font-size:9px;font-weight:750;cursor:pointer}
.extraBody{display:none;margin-top:12px;border-top:1px solid #2d2d2a}
.extraBody.open{display:block}
.extraBody .batchItem{border-color:#2b2b28;color:#aaa}
.extraBody .batchItem b{color:#fff}


.journey{
  margin-top:24px;border:1px solid var(--line);border-radius:16px;overflow:hidden;background:#fff;
}
.journeyHero{
  padding:24px 24px 20px;background:#111;color:#fff;
  display:grid;grid-template-columns:1fr auto;gap:22px;align-items:end
}
.journeyEyebrow{font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:#92928d;font-weight:800}
.journeyHero h2{margin:8px 0 5px;font-size:24px;letter-spacing:-.04em}
.journeyHero p{margin:0;color:#aaa9a3;font-size:10px;line-height:1.5;max-width:650px}
.journeyMoney{text-align:right}
.journeyMoney span{display:block;font-size:9px;color:#92928d;text-transform:uppercase;letter-spacing:.08em}
.journeyMoney b{display:block;font-size:27px;letter-spacing:-.04em;margin-top:4px}
.progressLine{height:4px;background:#292927}
.progressFill{height:100%;width:0;background:var(--accent);transition:width .25s ease}
.journeyBody{padding:0 24px}
.guideStep{
  display:grid;grid-template-columns:38px 1fr;gap:14px;padding:22px 0;border-bottom:1px solid var(--line)
}
.guideStep:last-child{border-bottom:0}
.guideCheck{
  width:32px;height:32px;border-radius:50%;border:1px solid var(--line2);background:#fafaf7;
  display:grid;place-items:center;font-size:11px;font-weight:800;cursor:pointer;transition:.16s ease
}
.guideStep.done .guideCheck{background:var(--accent);border-color:var(--accent);color:#152000}
.guideStep.done .guideTitle{text-decoration:line-through;text-decoration-color:#c8c8c0;color:#777}
.guideTop{display:flex;justify-content:space-between;align-items:flex-start;gap:14px}
.guideKicker{font-size:9px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;font-weight:800}
.guideTitle{font-size:16px;font-weight:800;letter-spacing:-.025em;margin-top:3px}
.guideDesc{font-size:10px;color:var(--muted);line-height:1.55;margin-top:6px;max-width:760px}
.guideTag{white-space:nowrap;border:1px solid var(--line);border-radius:999px;padding:5px 7px;font-size:8px;color:var(--muted);font-weight:800}
.guideInside{margin-top:13px}
.clickPath{
  display:flex;gap:6px;align-items:center;flex-wrap:wrap;padding:11px 12px;
  background:#fafaf7;border:1px solid var(--line);border-radius:10px;font-size:10px
}
.clickPath b{font-weight:800}
.chev{color:#aaa}
.guideButtons{display:flex;gap:8px;flex-wrap:wrap;margin-top:10px}
.guideBtn{
  border:0;background:#111;color:#fff;padding:9px 11px;border-radius:8px;
  font-size:9px;font-weight:800;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center
}
.guideBtn.secondary{background:#fff;color:#222;border:1px solid var(--line)}
.routeGrid{display:grid;grid-template-columns:1fr 1fr;gap:9px;margin-top:12px}
.routeCard{border:1px solid var(--line);border-radius:11px;padding:13px;background:#fff}
.routeCard.recommended{border-color:#c5da8e;background:#fbfff2}
.routeBadge{display:inline-block;font-size:8px;text-transform:uppercase;letter-spacing:.07em;font-weight:800;color:#546b1a;background:#eaffb9;border-radius:999px;padding:4px 6px;margin-bottom:7px}
.routeCard b{display:block;font-size:12px}
.routeCard p{margin:5px 0 0;color:var(--muted);font-size:9px;line-height:1.5}
.providerList{display:grid;gap:7px;margin-top:10px}
.providerRow{
  display:grid;grid-template-columns:1fr auto;gap:10px;align-items:center;
  padding:10px 11px;border:1px solid var(--line);border-radius:10px
}
.providerName{font-size:10px;font-weight:800}
.providerMeta{font-size:9px;color:var(--muted);margin-top:3px}
.providerAction{border:0;background:transparent;font-size:9px;font-weight:800;text-decoration:underline;text-underline-offset:3px;cursor:pointer}
.guideNote{
  margin-top:10px;padding:10px 11px;background:#f7f5eb;border-radius:9px;
  color:#71683e;font-size:9px;line-height:1.5
}
.guideAlert{
  margin-top:10px;padding:10px 11px;background:#fff5f2;border:1px solid #f2ded8;border-radius:9px;
  color:#8c493c;font-size:9px;line-height:1.5
}
.categoryGuide{display:flex;gap:6px;flex-wrap:wrap;margin-top:9px}
.categoryGuide span{font-size:8px;padding:5px 7px;border:1px solid var(--line);border-radius:999px;background:#fff}
.guideDoneMessage{display:none;margin:0 24px 22px;padding:14px;border-radius:11px;background:#efffcb;color:#334a00;font-size:10px;font-weight:750}
@media(max-width:820px){
  .journeyHero{grid-template-columns:1fr}
  .journeyMoney{text-align:left}
  .journeyBody{padding:0 18px}
  .journeyHero{padding:21px 18px}
  .routeGrid{grid-template-columns:1fr}
  .providerRow{grid-template-columns:1fr}
  .guideTop{flex-direction:column}
}

.packetRow{display:flex;align-items:center;justify-content:space-between;gap:18px;padding:22px 0 0;margin-top:18px;border-top:1px solid var(--line)}
.packetRow b{font-size:13px}
.packetRow p{margin:4px 0 0;color:var(--muted);font-size:10px}
.divider{height:1px;background:var(--line);margin:24px 0 0}

.detailsCard{margin-top:0}
.detailsSummary{list-style:none;cursor:pointer;padding:18px 0;color:#333}
.detailsSummary::-webkit-details-marker{display:none}
.detailsSummary b{font-size:12px}
.detailsSummary span{font-size:10px;color:var(--muted)}
.detailsInner{padding:0 0 8px}
.tableActions{display:flex;align-items:center;justify-content:space-between;gap:14px;margin-bottom:8px}
.tableActions .left{font-size:9px;color:var(--muted)}
table{width:100%;border-collapse:collapse}
th,td{padding:9px 7px;border-bottom:1px solid var(--line);text-align:left;font-size:10px}
th{color:var(--muted);font-size:8px;text-transform:uppercase;letter-spacing:.06em;font-weight:750}
.status{display:inline-flex;padding:4px 6px;border-radius:999px;font-size:8px;font-weight:800;white-space:nowrap}
.high{background:var(--greenSoft);color:var(--green)}
.verify{background:var(--amberSoft);color:var(--amber)}
.amt{font-weight:760;white-space:nowrap}
.merchant{font-weight:650}
.yearhead td{background:#fafaf7;font-weight:800}
.disclaimer{margin-top:18px;padding-top:17px;border-top:1px solid var(--line);color:var(--muted);font-size:9px;line-height:1.5}
.footer{margin-top:10px;color:#a1a19a;font-size:8px}
.toast{display:none;position:fixed;right:20px;bottom:20px;background:#111;color:#fff;padding:10px 12px;border-radius:9px;font-size:10px;box-shadow:0 12px 30px rgba(0,0,0,.14)}

@media(max-width:820px){
  .wrap{padding:16px 15px 48px}
  .secure{font-size:0}.secureDot{margin:0}
  .hero{grid-template-columns:1fr;gap:28px;padding:45px 0 32px}
  h1{font-size:47px}
  .heroText{font-size:15px}
  .previewMoney{font-size:41px}
  .resultGrid{grid-template-columns:1fr}
  .fastBox{border-left:0;border-top:1px solid var(--line);padding:18px 0 0}
  .refund{font-size:51px}
  .resultTop,.resultBody{padding:22px 18px}
  .actionCard{grid-template-columns:32px 1fr}
  .actionMoney{grid-column:2}
  .qrow{grid-template-columns:1fr}
  .uploaderHead{flex-direction:column}
  .formatPills{justify-content:flex-start}
  .upload{align-items:flex-start}
  .uploadPick{display:none}
  .summaryNumbers{grid-template-columns:1fr}
  .summaryNumber+ .summaryNumber{border-left:0;border-top:1px solid var(--line)}
  .paywall{grid-template-columns:1fr}
  .paywallPrice{justify-content:space-between}
  .paySecure{text-align:left}
  .packetRow{align-items:flex-start;flex-direction:column}
  .sectionTitleRow{align-items:flex-start;flex-direction:column}
}
</style>
</head>
<body>
<div class="wrap">
  <header class="header">
    <div class="brand">
    <svg class="brandLogo" viewBox="0 0 44 44" xmlns="http://www.w3.org/2000/svg" aria-hidden="true">
      <rect x="1" y="1" width="42" height="42" rx="13" fill="#B7FF2A"/>
      <path d="M15 12.5h10l5 5v13.5H15z" fill="none" stroke="#172000" stroke-width="2.4" stroke-linejoin="round"/>
      <path d="M25 12.5v5h5" fill="none" stroke="#172000" stroke-width="2.4" stroke-linejoin="round"/>
      <path d="m18.5 25 3 3 6-7" fill="none" stroke="#172000" stroke-width="2.8" stroke-linecap="round" stroke-linejoin="round"/>
    </svg>
    <span class="brandName">СделатьВычет</span>
    <span class="beta">beta</span>
  </div>
    <div class="secure"><span class="secureDot"></span>Файл не сохраняется после анализа</div>
  </header>

  <section class="hero">
    <div class="heroCopy">
      <div class="eyebrow"><span class="eyebrowDot"></span>Поиск вычетов по банковской выписке</div>
      <h1>Найдём ваш налоговый вычет за 5 минут.</h1>
      <p class="heroText">Загрузите банковскую выписку. СделатьВычет поможет найти расходы, которые могут дать налоговый вычет, оценит сумму возврата и сведёт всё к нескольким понятным действиям.</p>
      <div class="publicCounter" id="publicCounter"><span>Уже нашли вычетов на</span><b id="publicCounterValue">— ₽</b></div>
      <div class="heroActions">
        <label for="file" class="btn btnPrimary">Загрузить выписку <span>→</span></label>
        <span class="heroHint">Бесплатный поиск → детали за 499 ₽</span>
      </div>
    </div>
    <div class="preview">
      <div class="previewTop"><span>Результат</span><span class="previewPill">SDELATVYCHET 2.4.2</span></div>
      <div class="previewLabel">Потенциальный вычет</div>
      <div class="previewMoney">от 20 208 ₽</div>
      <div class="previewFast"><i>✓</i>Сразу покажем, где искать вычет</div>
      <div class="previewSteps">
        <div class="previewStep"><b>1. Получить справки</b>Мы сами сгруппируем организации и подготовим запросы.</div>
        <div class="previewStep"><b>2. Ответить на вопросы</b>Уточнить только те операции, где без вас нельзя.</div>
      </div>
    </div>
  </section>

  <section class="proofStrip">
    <div class="proofCard">
      <div class="proofNum">1</div>
      <div><b>Загрузите выписку</b><span>PDF, CSV или XLSX. Поддерживаем крупнейшие банки РФ.</span></div>
    </div>
    <div class="proofCard">
      <div class="proofNum">2</div>
      <div><b>Бесплатно увидьте ориентир</b><span>Сначала покажем найденные расходы и оценку возможного вычета.</span></div>
    </div>
    <div class="proofCard">
      <div class="proofNum">3</div>
      <div><b>Откройте готовый план</b><span>После оплаты получите понятный отчёт, тексты запросов и маршрут до подачи.</span></div>
    </div>
  </section>

  <section class="uploaderCard">
    <div class="uploaderHead">
      <div><h2>Проверить выписку</h2><p>PDF крупнейших банков распознаются автоматически. Для CSV/XLSX нужны дата, описание операции и сумма.</p></div>
      <div style="display:flex;gap:8px;flex-wrap:wrap;margin:10px 0 0"><span class="pill">До оплаты — бесплатно</span><span class="pill">Отчёт за 499 ₽</span><span class="pill">Данные не сохраняются как файл</span></div><div class="formatPills"><span class="formatPill">PDF</span><span class="formatPill">CSV</span><span class="formatPill">XLSX</span></div>
    </div>
    <label class="upload" id="drop">
      <input id="file" type="file" accept=".pdf,.csv,.xlsx,.xlsm"/>
      <div class="uploadLeft">
        <div class="uploadIcon">↥</div>
        <div class="uploadMeta">
          <div class="uploadTitle" id="fname">Загрузите банковскую выписку</div>
          <div class="uploadSub" id="fileSub">Поддержка PDF крупнейших банков РФ · CSV / XLSX · безопасная загрузка</div>
        </div>
      </div>
      <span class="uploadPick" id="uploadPick">Выбрать файл</span>
    </label>

    <div class="legalConsent">
      <label class="legalRow">
        <input type="checkbox" id="termsAccept">
        <span>Я принимаю <a href="/terms" target="_blank">Пользовательское соглашение</a>
        (версия 2026-08-27-v1).</span>
      </label>
      <label class="legalRow">
        <input type="checkbox" id="pdConsent">
        <span>Я отдельно даю <a href="/consent" target="_blank">Согласие на обработку персональных данных</a>
        и ознакомился с <a href="/privacy" target="_blank">Политикой обработки персональных данных</a>.</span>
      </label>
      <div class="legalMini">Загружайте только банковскую выписку. Не загружайте медицинские документы, диагнозы или паспортные сканы.</div>
    </div>

    <div class="uploadBottom">
      <button class="btn btnBrand" id="analyze" disabled>Найти мой вычет</button>
      <span class="parser" id="parserHint">Файл передаётся по HTTPS и не сохраняется приложением</span>
    </div>
    <div class="loader" id="loader">
      <div class="loaderTop"><b>Анализируем операции</b><span>ищем медицину, фитнес, обучение, страхование…</span></div>
      <div class="track"><div class="bar"></div></div>
    </div>
    <div id="error"></div>
  </section>

  <section class="summary" id="summary">
    <div class="summaryCard">
      <div class="summaryTop">
        <div class="summaryBadge"><i>✓</i>Анализ готов</div>
        <div class="summaryTitle">Мы нашли расходы, которые могут дать вам вычет</div>
        <div class="summarySub">Сначала бесплатно покажем найденную сумму. После оплаты откроем детали и готовый план действий.</div>
        <div class="summaryNumbers">
          <div class="summaryNumber"><span>Нашли подходящих расходов</span><b id="summaryExpenses">—</b></div>
          <div class="summaryNumber"><span>Потенциальный возврат</span><b id="summaryRefund">—</b></div>
        </div>
        <div class="paymentStatus" id="paymentStatus"></div>
      </div>
      <div class="paywall">
        <div>
          <div class="paywallLabel">Полный отчёт</div>
          <div class="paywallTitle">Откроем, где именно лежат ваши деньги</div>
          <div class="paywallText">Покажем найденные операции, соберём организации по категориям, подготовим тексты запросов и дадим пошаговый маршрут до подачи вычета.</div>
        </div>
        <div>
          <div class="paywallPrice"><div class="price">499 ₽</div><button class="payBtn" id="payBtn">Получить отчёт</button></div>
          <div class="paySecure">Разовая оплата · ЮKassa</div>
        </div>
      </div>
    </div>
  </section>

  <section id="results">
    <div class="successBanner" id="successBanner">
      <b>Оплата прошла успешно — полный отчёт открыт</b>
      <p>Теперь вы видите найденные операции, категории, организации и пошаговый план до подачи вычета.</p>
    </div>
    <div class="resultShell">
      <div class="resultTop">
        <div class="unlockBadge">✓ Полный отчёт оплачен и открыт</div>
        <div class="resultBadge"><i>✓</i>Ваш результат</div>
        <div class="resultGrid">
          <div>
            <div class="resultLabel">Потенциальный возврат</div>
            <div class="refund" id="refund"><span class="from">от</span>—</div>
            <div class="resultMeta" id="resultMeta"></div>
            <div class="metricTiles">
              <div class="metricTile"><div class="label">Найдено расходов</div><div class="value" id="tileExpenses">—</div><div class="meta">Сумма операций, которые сервис отнёс к возможному вычету</div></div>
              <div class="metricTile"><div class="label">Быстрый путь</div><div class="value" id="tileFast">—</div><div class="meta">Сколько можно вернуть по простому сценарию</div></div>
              <div class="metricTile"><div class="label">Доп. потенциал</div><div class="value" id="tileExtra">—</div><div class="meta">Столько ещё может добавиться при доп. подтверждениях</div></div>
            </div>
          </div>
          <div class="fastBox">
            <div class="fastLabel">Быстрый путь</div>
            <div class="fastMoney" id="fastRefund">—</div>
            <div class="fastText" id="fastText"></div>
          </div>
        </div>
      </div>

      <div class="resultBody">
        <div class="sectionTitleRow">
          <div><h2>Что СделатьВычет подготовил</h2><p id="actionLead">Готовые запросы и уточнения по найденным расходам.</p></div>
          <div class="tiny" id="foundLabel"></div>
        </div>
        <div class="groups" id="groups"></div>

        <section class="journey" id="journey">
          <div class="journeyHero">
            <div>
              <div class="journeyEyebrow">Ваш план возврата</div>
              <h2>Как получить деньги шаг за шагом</h2>
              <p id="journeyIntro">Мы превратили найденные расходы в три понятных шага. Начните с организаций, у которых нужно получить подтверждения.</p>
            </div>
            <div class="journeyMoney"><span>Ваш ориентир</span><b id="journeyRefund">—</b></div>
          </div>
          <div class="progressLine"><div class="progressFill" id="guideProgress"></div></div>
          <div class="journeyBody" id="journeyBody"></div>
          <div class="guideDoneMessage" id="guideDoneMessage">✓ Всё готово. Теперь осталось дождаться результата рассмотрения в ФНС.</div>
        </section>

        <div class="sectionTitleRow" style="margin-top:28px">
          <div><h2>Подробности отчёта</h2><p>Здесь можно проверить операции, подготовить запросы и скорректировать расчёт.</p></div>
        </div>
        <div class="actions" id="actions"></div>

        <div class="extra" id="extra">
          <div class="extraHead">
            <div><div class="extraMoney" id="extraMoney"></div><div class="extraText" id="extraText"></div></div>
            <button class="extraBtn" id="extraToggle">Разобраться</button>
          </div>
          <div class="extraBody" id="extraBody"></div>
        </div>

        <div class="packetRow">
          <div><b>Готовый персональный пакет</b><p>Список выбранных расходов + суммы + короткий чек-лист для получения вычета.</p></div>
          <button class="btn" id="packet">Скачать пакет</button>
        </div>

        <div class="divider"></div>

        <details class="detailsCard">
          <summary class="detailsSummary"><div><b>Все найденные операции</b><br><span>Нужны только для проверки и ручной корректировки результата.</span></div></summary>
          <div class="detailsInner">
            <div class="tableActions"><div class="left">Жёлтые операции включены по умолчанию. Снимите галочку, если операция не подходит.</div><button class="btn btnGhost" id="selectAll">Снять все</button></div>
            <div style="overflow:auto">
              <table><thead><tr><th>Включить</th><th>Дата</th><th>Категория</th><th>Организация</th><th>Сумма</th><th>Статус</th></tr></thead><tbody id="tbody"></tbody></table>
            </div>
          </div>
        </details>

        <div class="disclaimer">Расчёт является ориентиром: СделатьВычет использует базовые 13% и консервативный общий социальный лимит 150 000 ₽ на год. Фактическое право на вычет и сумма зависят от подтверждающих документов и уплаченного НДФЛ.</div>
        <div class="footer" id="footer"></div>
      </div>
    </div>
  </section>

  <footer class="legalFooter">
    <button type="button" id="analyticsSettingsBtn" style="border:0;background:none;padding:0;color:inherit;text-decoration:underline;cursor:pointer;font:inherit">Настройки аналитики</button>
    <a href="/terms">Пользовательское соглашение</a>
    <a href="/privacy">Персональные данные</a>
    <a href="/consent">Согласие на обработку ПДн</a>
    <span>СделатьВычет · оператор: Колосов Роман Михайлович · НПД · ИНН 772072450119</span>
    <a href="mailto:inbox@sdelatvychet.ru">inbox@sdelatvychet.ru</a>
  </footer>
</div>
<div class="analyticsConsent" id="analyticsConsent">
  <b>Помочь нам улучшать СделатьВычет?</b>
  <p>При разрешении подключим Яндекс Метрику без Вебвизора. Содержимое выписки, названия операций и точные суммы в Метрику не передаются.</p>
  <div class="analyticsConsentActions"><button class="allow" id="allowAnalytics">Разрешить аналитику</button><button class="deny" id="denyAnalytics">Только необходимые</button></div>
</div>
<div class="mobileSticky" id="mobileSticky">
  <div><span>СделатьВычет</span><b id="stickyText">Проверить выписку</b></div>
  <button id="stickyBtn" type="button">Начать →</button>
</div>
<div class="toast" id="toast"></div>
<script>
const METRIKA_COUNTER_ID=Number('__METRIKA_COUNTER_ID__')||0;
const ANALYTICS_SESSION_KEY='sv_analytics_session';
const ANALYTICS_CONSENT_KEY='sv_analytics_consent';

function makeAnalyticsSession(){
  let id=localStorage.getItem(ANALYTICS_SESSION_KEY);
  if(!id){
    try{id=crypto.randomUUID().replaceAll('-','')}catch(e){id='sv'+Math.random().toString(36).slice(2)+Date.now().toString(36)}
    localStorage.setItem(ANALYTICS_SESSION_KEY,id);
  }
  return id;
}
const analyticsSessionId=makeAnalyticsSession();

const publicRub=n=>new Intl.NumberFormat('ru-RU',{maximumFractionDigits:0}).format(Math.round(Number(n)||0))+' ₽';
async function refreshPublicCounter(){
  try{
    const r=await fetch('/api/public-stats',{cache:'no-store'});
    const d=await r.json();
    if(r.ok&&document.getElementById('publicCounterValue')){
      document.getElementById('publicCounterValue').textContent=publicRub(d.found_refund);
    }
  }catch(e){}
}


function analyticsAttribution(){
  const q=new URLSearchParams(location.search);
  let ref='';
  try{ref=document.referrer?new URL(document.referrer).hostname:''}catch(e){}
  const stored=JSON.parse(sessionStorage.getItem('sv_attribution')||'{}');
  const a={
    utm_source:q.get('utm_source')||stored.utm_source||'',
    utm_medium:q.get('utm_medium')||stored.utm_medium||'',
    utm_campaign:q.get('utm_campaign')||stored.utm_campaign||'',
    utm_content:q.get('utm_content')||stored.utm_content||'',
    utm_term:q.get('utm_term')||stored.utm_term||'',
    referrer_host:stored.referrer_host||ref||''
  };
  sessionStorage.setItem('sv_attribution',JSON.stringify(a));
  return a;
}
const attribution=analyticsAttribution();

function sensitiveReturnUrl(){
  const q=new URLSearchParams(location.search);
  return q.has('analysis')||q.has('payment')||q.has('paid');
}
function metrikaAllowed(){return localStorage.getItem(ANALYTICS_CONSENT_KEY)==='yes'}
function sanitizedMetrikaUrl(){
  return location.origin+location.pathname;
}
function loadMetrika(){
  if(!METRIKA_COUNTER_ID||!metrikaAllowed()||window.__svMetrikaLoaded)return;
  window.__svMetrikaLoaded=true;
  window.ym=window.ym||function(){(window.ym.a=window.ym.a||[]).push(arguments)};
  window.ym.l=1*new Date();
  const s=document.createElement('script');
  s.async=true;
  s.src='https://mc.yandex.ru/metrika/tag.js?id='+METRIKA_COUNTER_ID;
  document.head.appendChild(s);
  ym(METRIKA_COUNTER_ID,'init',{
    ssr:true,
    webvisor:false,
    clickmap:false,
    trackLinks:false,
    accurateTrackBounce:true,
    sendTitle:false,
    url:sanitizedMetrikaUrl()
  });
}
function metrikaGoal(event){
  if(METRIKA_COUNTER_ID&&metrikaAllowed()&&typeof window.ym==='function'){
    try{ym(METRIKA_COUNTER_ID,'reachGoal',event)}catch(e){}
  }
}
window.sdelatVychetAnalyticsStatus=function(){
  return {
    counterId:METRIKA_COUNTER_ID,
    consent:localStorage.getItem(ANALYTICS_CONSENT_KEY),
    metrikaLoaded:!!window.__svMetrikaLoaded,
    ymAvailable:typeof window.ym==='function',
    sanitizedUrl:sanitizedMetrikaUrl()
  };
};

function track(event){
  const body={event,session_id:analyticsSessionId,...attribution};
  try{
    fetch('/api/analytics/event',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body),keepalive:true}).catch(()=>{});
  }catch(e){}
  metrikaGoal(event);
}
function initAnalyticsConsent(){
  const box=document.getElementById('analyticsConsent');
  if(!box)return;
  const settingsBtn=document.getElementById('analyticsSettingsBtn');
  if(settingsBtn)settingsBtn.onclick=()=>{
    localStorage.removeItem(ANALYTICS_CONSENT_KEY);
    box.classList.add('show');
  };
  const choice=localStorage.getItem(ANALYTICS_CONSENT_KEY);
  if(METRIKA_COUNTER_ID&&!choice)box.classList.add('show');
  if(choice==='yes')loadMetrika();
  document.getElementById('allowAnalytics').onclick=()=>{localStorage.setItem(ANALYTICS_CONSENT_KEY,'yes');box.classList.remove('show');loadMetrika()};
  document.getElementById('denyAnalytics').onclick=()=>{localStorage.setItem(ANALYTICS_CONSENT_KEY,'no');box.classList.remove('show')};
}

let chosenFile=null,result=null,analysisId=null;
const TERMS_VERSION='2026-08-27-v1',CONSENT_VERSION='2026-08-27-v1';
const file=document.getElementById('file'),drop=document.getElementById('drop'),analyze=document.getElementById('analyze');
const termsAccept=document.getElementById('termsAccept'),pdConsent=document.getElementById('pdConsent');
function refreshAnalyzeButton(){analyze.disabled=!(chosenFile&&termsAccept.checked&&pdConsent.checked)}
termsAccept.onchange=refreshAnalyzeButton;pdConsent.onchange=refreshAnalyzeButton;
file.onchange=()=>setFile(file.files[0]);
drop.addEventListener('click',()=>track('upload_click'));
['dragover','dragenter'].forEach(e=>drop.addEventListener(e,x=>{x.preventDefault();drop.classList.add('drag')}));
['dragleave','drop'].forEach(e=>drop.addEventListener(e,x=>{x.preventDefault();drop.classList.remove('drag')}));
drop.addEventListener('drop',e=>{if(e.dataTransfer.files[0])setFile(e.dataTransfer.files[0])});
function setFile(f){
  if(!f)return;
  chosenFile=f;
  track('file_selected');
  drop.classList.add('selected');
  document.getElementById('fname').textContent=f.name;
  document.getElementById('fileSub').textContent=(f.size/1024/1024).toFixed(1)+' МБ · готово к анализу';
  document.getElementById('uploadPick').textContent='Заменить';
  refreshAnalyzeButton();
  document.getElementById('parserHint').textContent='После анализа исходный файл не сохраняется';
}
const rub=n=>new Intl.NumberFormat('ru-RU',{maximumFractionDigits:0}).format(n)+' ₽';
function escapeHtml(s){return String(s).replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}[m]))}
function ids(){return [...document.querySelectorAll('input[data-id]:checked')].map(x=>x.dataset.id)}
function selected(){const set=new Set(ids());return result.candidates.filter(c=>set.has(c.id))}
function uniq(items){const a=[];items.forEach(c=>{const m=(c.merchant||'').trim();if(m&&m!=='Не удалось определить'&&!a.includes(m))a.push(m)});return a}
function totalCandidatesAmount(items){return (items||[]).reduce((s,x)=>s+(Number(x.amount)||0),0)}

function fastPathItems(items){
  return (items||[]).filter(c=>['medicine','fitness','education','insurance'].includes(c.category));
}
function fastPathRefund(items){
  return subsetRefund(fastPathItems(items));
}

function subsetRefund(items){const by={};items.forEach(c=>(by[c.year]??=0,by[c.year]+=c.amount));return Object.values(by).reduce((s,v)=>s+Math.min(v,150000)*.13,0)}
function toast(s){const e=document.getElementById('toast');e.textContent=s;e.style.display='block';setTimeout(()=>e.style.display='none',1800)}
function copyText(s){navigator.clipboard?.writeText(s).then(()=>toast('Готовый запрос скопирован')).catch(()=>toast('Скопируйте текст вручную'))}

analyze.onclick=async()=>{
  track('analysis_started');
  document.getElementById('error').style.display='none';
  document.getElementById('loader').style.display='block';
  analyze.disabled=true;
  const fd=new FormData();
  fd.append('file',chosenFile);
  fd.append('terms_accepted',termsAccept.checked?'1':'0');
  fd.append('pd_consent',pdConsent.checked?'1':'0');
  fd.append('terms_version',TERMS_VERSION);
  fd.append('consent_version',CONSENT_VERSION);
  fd.append('analytics_session_id',analyticsSessionId);
  fd.append('utm_source',attribution.utm_source);
  fd.append('utm_medium',attribution.utm_medium);
  fd.append('utm_campaign',attribution.utm_campaign);
  fd.append('utm_content',attribution.utm_content);
  fd.append('utm_term',attribution.utm_term);
  fd.append('referrer_host',attribution.referrer_host);
  try{
    const r=await fetch('/api/analyze',{method:'POST',body:fd});
    const data=await r.json();
    if(!r.ok)throw new Error(data.detail||'Ошибка анализа');
    analysisId=data.analysis_id;
    localStorage.setItem('sdelatVychetAnalysisId',analysisId);
    metrikaGoal('analysis_success');
    refreshPublicCounter();
    showSummary(data);
  }catch(e){
    track('analysis_error');
    const er=document.getElementById('error');er.textContent=e.message;er.style.display='block'
  }finally{
    document.getElementById('loader').style.display='none';refreshAnalyzeButton()
  }
};


function updateSticky(mode){
  const text=document.getElementById('stickyText'),btn=document.getElementById('stickyBtn');
  if(!text||!btn)return;
  if(mode==='summary'){text.textContent='Отчёт доступен за 499 ₽';btn.textContent='Открыть →';btn.onclick=()=>document.getElementById('summary').scrollIntoView({behavior:'smooth'});}
  else if(mode==='report'){text.textContent='Ваш план готов';btn.textContent='К плану →';btn.onclick=()=>document.getElementById('journey').scrollIntoView({behavior:'smooth'});}
  else{text.textContent='Проверить выписку';btn.textContent='Начать →';btn.onclick=()=>document.querySelector('.uploaderCard').scrollIntoView({behavior:'smooth'});}
}
updateSticky('start');

function showSummary(data){
  track('result_view');track('paywall_view');
  document.getElementById('results').style.display='none';
  document.getElementById('summary').style.display='block';
  document.getElementById('summaryExpenses').textContent=rub(data.expenses_found);
  document.getElementById('summaryRefund').textContent='от '+rub(data.refund_from);
  updateSticky('summary');
  document.getElementById('summary').scrollIntoView({behavior:'smooth',block:'start'});
}

async function buyReport(){
  if(!analysisId)return;
  track('payment_click');
  const btn=document.getElementById('payBtn');
  const status=document.getElementById('paymentStatus');
  btn.disabled=true;btn.textContent='Открываем оплату…';
  status.style.display='none';
  try{
    const r=await fetch('/api/create-payment',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({analysis_id:analysisId})
    });
    const d=await r.json();
    if(!r.ok)throw new Error(d.detail||'Не удалось создать оплату');
    if(d.status==='succeeded'){await unlockReport();return}
    window.location.href=d.confirmation_url;
  }catch(e){
    track('payment_error');
    status.textContent=e.message;
    status.style.display='block';
    btn.disabled=false;btn.textContent='Получить отчёт';
  }
}
document.getElementById('payBtn').onclick=buyReport;

async function unlockReport(){
  if(!analysisId)return;
  const status=document.getElementById('paymentStatus');
  status.style.display='block';status.textContent='Проверяем оплату…';
  try{
    const sr=await fetch('/api/payment-status?analysis_id='+encodeURIComponent(analysisId));
    const sd=await sr.json();
    if(!sr.ok)throw new Error(sd.detail||'Не удалось проверить оплату');
    if(!sd.paid){
      status.textContent=sd.status==='pending'?'Платёж ещё обрабатывается. Обновим статус через несколько секунд.':'Оплата не подтверждена.';
      if(sd.status==='pending')setTimeout(unlockReport,2500);
      return;
    }
    const rr=await fetch('/api/report/'+encodeURIComponent(analysisId));
    const report=await rr.json();
    if(!rr.ok)throw new Error(report.detail||'Не удалось открыть отчёт');
    result=report; result.paid_unlocked=true;
    loadMetrika();
    metrikaGoal('payment_success');
    metrikaGoal('report_view');
    document.getElementById('summary').style.display='none';
    render();
    const sb=document.getElementById('successBanner'); if(sb) sb.classList.add('show');
    const url=new URL(window.location.href);
    url.searchParams.delete('payment');url.searchParams.delete('paid');
    url.searchParams.set('analysis',analysisId);
    history.replaceState({},'',url);
  }catch(e){
    status.textContent=e.message;
  }
}

async function resumeAfterPayment(){
  const params=new URLSearchParams(window.location.search);
  const fromUrl=params.get('analysis');
  analysisId=fromUrl||localStorage.getItem('sdelatVychetAnalysisId');
  if(!analysisId)return;
  if(params.get('payment')==='return'||params.get('paid')==='1'){
    document.getElementById('summary').style.display='block';
    document.getElementById('summaryExpenses').textContent='—';
    document.getElementById('summaryRefund').textContent='—';
    await unlockReport();
  }
}
window.addEventListener('DOMContentLoaded',()=>{
  initAnalyticsConsent();
  track('visit');
  refreshPublicCounter();
  resumeAfterPayment();
});


const FNS_LK_URL='https://lkfl2.nalog.ru/lkfl/';
const GUIDE_STORAGE_PREFIX='sdelatVychetGuide:';

function guideStorageKey(){
  return GUIDE_STORAGE_PREFIX+(analysisId||'demo');
}
function getGuideState(){
  try{return JSON.parse(localStorage.getItem(guideStorageKey())||'{}')}catch(e){return {}}
}
function setGuideState(state){
  localStorage.setItem(guideStorageKey(),JSON.stringify(state));
}
function toggleGuideStep(step){
  const st=getGuideState();st[step]=!st[step];setGuideState(st);
  if(st[step])track('guide_step_'+step);
  if([1,2,3].every(x=>st[x]))track('guide_complete');
  renderJourney();
}
function guideCompletedCount(){
  const st=getGuideState();return [1,2,3].filter(x=>st[x]).length
}
function providerGroups(items){
  const by={};
  items.forEach(c=>{
    const key=c.year+'||'+c.category+'||'+c.merchant;
    (by[key]??=[]).push(c)
  });
  return Object.values(by);
}
function categoryDoc(category){
  const docs={
    medicine:'справку об оплате медицинских услуг (КНД 1151156)',
    fitness:'справку об оплате физкультурно‑оздоровительных услуг (КНД 1151160)',
    education:'справку об оплате образовательных услуг (КНД 1151158)',
    insurance:'справку об уплате страховых взносов (КНД 1151159)'
  };
  return docs[category]||'подтверждающий документ';
}
function requestTextFor(arr){
  const c=arr[0],year=c.year,dates=arr.map(x=>x.date+' — '+rub(x.amount)).join('; ');
  if(c.category==='medicine')return `Здравствуйте! Прошу направить в ФНС сведения о моих расходах на медицинские услуги за ${year} год для получения социального налогового вычета. Если электронная передача в ФНС невозможна, прошу выдать справку об оплате медицинских услуг для налогового органа (КНД 1151156). Найденные оплаты: ${dates}.`;
  if(c.category==='fitness')return `Здравствуйте! Прошу направить в ФНС сведения о моих расходах на физкультурно-оздоровительные услуги за ${year} год для получения социального налогового вычета. Если электронная передача невозможна, прошу выдать справку для налогового органа (КНД 1151160). Найденные оплаты: ${dates}.`;
  if(c.category==='education')return `Здравствуйте! Прошу направить в ФНС сведения о расходах на обучение за ${year} год для получения социального налогового вычета. Если электронная передача невозможна, прошу выдать справку для налогового органа (КНД 1151158). Найденные оплаты: ${dates}.`;
  if(c.category==='insurance')return `Здравствуйте! Прошу сообщить вид договора и, если он даёт право на социальный налоговый вычет, направить в ФНС сведения об уплаченных страховых взносах за ${year} год. Если электронная передача невозможна, прошу выдать справку для налогового органа (КНД 1151159). Найденные оплаты: ${dates}.`;
  return `Здравствуйте! Прошу предоставить документы, подтверждающие расходы за ${year} год для налогового вычета. Найденные оплаты: ${dates}.`;
}
function buildProviderRows(items){
  const groups=providerGroups(items);
  if(!groups.length)return '<div class="guideNote">По этим категориям отдельные справки не найдены.</div>';
  return `<div class="providerList">${groups.map(arr=>{
    const c=arr[0],sum=arr.reduce((s,x)=>s+x.amount,0),txt=requestTextFor(arr);
    return `<div class="providerRow">
      <div><div class="providerName">${c.emoji} ${escapeHtml(c.merchant)}</div><div class="providerMeta">${c.year} · ${arr.length} ${arr.length===1?'оплата':'оплат'} · ${rub(sum)} · ${escapeHtml(categoryDoc(c.category))}</div></div>
      <button class="providerAction" onclick='copyText(${JSON.stringify(txt)})'>Скопировать запрос</button>
    </div>`
  }).join('')}</div>`
}
function buildMedicinePharmacyNote(pharmacy){
  if(!pharmacy.length)return '';
  return `<div class="guideAlert"><b>Лекарства — отдельный путь.</b> По ${pharmacy.length} найденным аптечным покупкам одной справки от аптеки недостаточно. Сначала проверьте, есть ли назначение врача: рецепт или сведения из выписного эпикриза. Затем сохраните подтверждение оплаты.</div>`;
}
function renderJourney(){
  if(!result)return;
  const s=selected(),st=getGuideState();
  const serviceDocs=s.filter(c=>['medicine','fitness','education','insurance'].includes(c.category));
  const pharmacy=s.filter(c=>c.category==='pharmacy');
  const years=[...new Set(s.map(c=>c.year).filter(Boolean))].sort();
  const cats=[...new Set(s.map(c=>c.category_name))];
  const providers=providerGroups(serviceDocs);

  const steps=[];

  steps.push(`
    <div class="guideStep ${st[1]?'done':''}">
      <button class="guideCheck" onclick="toggleGuideStep(1)">${st[1]?'✓':'1'}</button>
      <div>
        <div class="guideTop">
          <div>
            <div class="guideKicker">Шаг 1 · начните отсюда</div>
            <div class="guideTitle">Получите документы по найденным расходам</div>
          </div>
          <span class="guideTag">${providers.length} организаций</span>
        </div>

        <div class="guideDesc">
          СделатьВычет уже нашёл подходящие платежи. Ниже — конкретные организации,
          суммы и готовые запросы. Вам не нужно искать, кому и что писать.
        </div>

        <div class="categoryGuide">${cats.map(x=>`<span>${escapeHtml(x)}</span>`).join('')}</div>

        <div class="guideInside">
          ${buildProviderRows(serviceDocs)}
          ${buildMedicinePharmacyNote(pharmacy)}

          <div class="guideNote">
            <b>Что просить у организации:</b> попросите передать сведения о расходах
            в ФНС для налогового вычета. Если организация не может передать их
            электронно — попросите справку для налогового органа. Готовый текст
            уже сформирован для каждой организации выше.
          </div>

          <div class="guideButtons">
            <button class="guideBtn secondary" onclick="toggleGuideStep(1)">Документы запросил</button>
          </div>
        </div>
      </div>
    </div>`);

  steps.push(`
    <div class="guideStep ${st[2]?'done':''}">
      <button class="guideCheck" onclick="toggleGuideStep(2)">${st[2]?'✓':'2'}</button>
      <div>
        <div class="guideTop">
          <div>
            <div class="guideKicker">Шаг 2 · когда получили подтверждения</div>
            <div class="guideTitle">Подайте вычет в Личном кабинете ФНС</div>
          </div>
          <span class="guideTag">${years.join(', ')||'нужный год'}</span>
        </div>

        <div class="guideDesc">
          Когда справки готовы или организации подтвердили передачу сведений,
          переходите в Личный кабинет ФНС. Дальше есть два сценария — выбирайте
          тот, который видите у себя.
        </div>

        <div class="routeGrid">
          <div class="routeCard recommended">
            <span class="routeBadge">если ФНС уже получила сведения</span>
            <b>Упрощённый вычет</b>
            <p>
              Откройте раздел «Вычеты». Если ФНС уже получила данные от клиники,
              фитнеса, страховой или другой организации, появится готовое или
              предзаполненное заявление. Проверьте суммы и подтвердите его.
            </p>
          </div>

          <div class="routeCard">
            <b>Если готового заявления нет</b>
            <p>
              Откройте «Декларации → Подать декларацию», выберите нужный год и
              внесите данные из полученных справок в раздел социальных вычетов.
              Для лекарств приложите назначение врача и подтверждение покупки.
            </p>
          </div>
        </div>

        <div class="guideInside">
          <div class="clickPath">
            <b>ЛК ФНС</b><span class="chev">→</span><b>Вычеты</b>
            <span class="chev">или</span><b>Декларации → Подать декларацию</b>
          </div>

          <div class="guideButtons">
            <a class="guideBtn" href="${FNS_LK_URL}" target="_blank" rel="noopener">Открыть ЛК ФНС ↗</a>
            <button class="guideBtn secondary" onclick="toggleGuideStep(2)">Вычет подан</button>
          </div>
        </div>
      </div>
    </div>`);

  steps.push(`
    <div class="guideStep ${st[3]?'done':''}">
      <button class="guideCheck" onclick="toggleGuideStep(3)">${st[3]?'✓':'3'}</button>
      <div>
        <div class="guideTop">
          <div>
            <div class="guideKicker">Шаг 3 · готово</div>
            <div class="guideTitle">Дождитесь результата от ФНС</div>
          </div>
          <span class="guideTag">финальный шаг</span>
        </div>

        <div class="guideDesc">
          После отправки ничего дополнительно делать не нужно. Следите за статусом
          заявления или декларации в Личном кабинете. Если ФНС попросит уточнение,
          оно появится там же.
        </div>

        <div class="guideInside">
          <div class="clickPath">
            <b>ЛК ФНС</b><span class="chev">→</span><b>Вычеты / Декларации</b>
            <span class="chev">→</span><b>Статус</b>
          </div>

          <div class="guideButtons">
            <a class="guideBtn secondary" href="${FNS_LK_URL}" target="_blank" rel="noopener">Проверить статус ↗</a>
            <button class="guideBtn secondary" onclick="toggleGuideStep(3)">Готово</button>
          </div>
        </div>
      </div>
    </div>`);

  document.getElementById('journeyBody').innerHTML=steps.join('');

  const done=[1,2,3].filter(x=>st[x]).length;
  document.getElementById('guideProgress').style.width=(done/3*100)+'%';
  document.getElementById('guideDoneMessage').style.display=done===3?'block':'none';
  document.getElementById('guideDoneMessage').textContent='✓ Всё сделано. Теперь остаётся дождаться результата рассмотрения в ФНС.';
  document.getElementById('journeyRefund').textContent='от '+rub(subsetRefund(s));
  document.getElementById('journeyIntro').textContent=`Мы нашли ${s.length} подходящих расходов. Чтобы получить вычет, пройдите три шага ниже — начинаем сразу с конкретных организаций и документов.`;
}

function buildBatch(items){
  const by={};items.forEach(c=>{const k=c.category+'||'+c.merchant;(by[k]??=[]).push(c)});let html='';
  Object.values(by).forEach(arr=>{const c=arr[0],sum=arr.reduce((s,x)=>s+x.amount,0);let req=c.category==='medicine'?'Прошу предоставить справку об оплате медицинских услуг для оформления налогового вычета.':c.category==='fitness'?'Прошу предоставить документы/сведения об оплате услуг для оформления спортивного налогового вычета.':'Прошу предоставить подтверждение оплаты услуг для оформления налогового вычета.';const txt=`Здравствуйте! ${req} Найденные оплаты: ${arr.map(x=>x.date+' — '+rub(x.amount)).join('; ')}.`;html+=`<div class="batchItem"><b>${escapeHtml(c.merchant)} · ${rub(sum)}</b>${arr.length} оплат · ${c.category_name}<div class="copy" onclick='copyText(${JSON.stringify(txt)})'>Скопировать готовый запрос →</div></div>`});return html
}

function renderActions(){
  const s=selected(),easy=s.filter(c=>['medicine','fitness','education'].includes(c.category)),questions=s.filter(c=>c.category==='insurance'),extras=s.filter(c=>['pharmacy','donation'].includes(c.category));
  const fast=[...easy,...questions],total=subsetRefund(s),fastR=fastPathRefund(s),extra=Math.max(0,total-fastR);let count=0,html='';
  if(easy.length){count++;const merchants=uniq(easy),easyR=subsetRefund(easy);html+=`<div class="actionCard"><div class="actionNo">${count}</div><div><div class="actionName">Получить справки одним пакетом</div><div class="actionDesc">${merchants.length} организаций, ${easy.length} найденных оплат. СделатьВычет уже сгруппировал, кому и что запросить.</div><button class="actionBtn" onclick="toggleBatch()">Подготовить все запросы</button><div class="batch" id="batch">${buildBatch(easy)}</div></div><div class="actionMoney">от ${rub(easyR)}</div></div>`}
  if(questions.length){count++;const merchants=uniq(questions),qR=subsetRefund(questions);let qhtml='';merchants.forEach(m=>{const related=questions.filter(c=>c.merchant===m);qhtml+=`<div class="qrow"><b>${escapeHtml(m)}</b><select onchange='insuranceAnswer(this,${JSON.stringify(related.map(x=>x.id))})'><option value="">Что это за страховка?</option><option value="keep">Жизнь / ДМС / подходящий договор</option><option value="drop">ОСАГО / каско / другое</option></select></div>`});html+=`<div class="actionCard"><div class="actionNo">${count}</div><div><div class="actionName">Ответить на ${merchants.length} коротких вопрос${merchants.length===1?'':'а'}</div><div class="actionDesc">Не нужно разбираться в договорах заранее — просто укажите тип найденной страховки.</div><button class="actionBtn" onclick="toggleQuestions()">Ответить</button><div class="questionBox" id="questionBox">${qhtml}</div></div><div class="actionMoney">до ${rub(qR)}</div></div>`}
  document.getElementById('actions').innerHTML=html||'<div class="actionCard"><div class="actionNo">✓</div><div><div class="actionName">Основной план уже готов</div><div class="actionDesc">Осталось проверить дополнительные расходы или скачать пакет.</div></div></div>';
  document.getElementById('tileExpenses').textContent=rub(totalCandidatesAmount(s));
  document.getElementById('tileFast').textContent='до '+rub(fastR);
  document.getElementById('tileExtra').textContent='до '+rub(extra);
  document.getElementById('refund').innerHTML='<span class="from">от</span>'+rub(total);document.getElementById('fastRefund').textContent='от '+rub(fastR);document.getElementById('fastText').textContent=count?`${fastR?Math.round(fastR/Math.max(total,1)*100):0}% найденного возврата — за ${count} ${count===1?'действие':'действия'}.`:'Сначала проверьте найденные операции.';document.getElementById('actionLead').textContent=count?`Вместо ${s.length} отдельных операций — всего ${count} ${count===1?'действие':'действия'}.`:'Сложные расходы вынесены отдельно.';
  const ex=document.getElementById('extra');if(extras.length){ex.style.display='block';document.getElementById('extraMoney').textContent='Ещё до '+rub(extra);const cats=[...new Set(extras.map(c=>c.category_name))];document.getElementById('extraText').textContent=`${cats.join(', ')} — ${extras.length} операций. Здесь нужно больше ручных подтверждений, поэтому они не мешают основному сценарию.`;document.getElementById('extraBody').innerHTML=extras.map(c=>`<div class="batchItem"><b>${c.emoji} ${escapeHtml(c.merchant)} · ${rub(c.amount)}</b>${c.date} · ${escapeHtml(c.note)}</div>`).join('')}else ex.style.display='none';
  renderJourney();
}
function toggleBatch(){const e=document.getElementById('batch');if(e)e.style.display=e.style.display==='block'?'none':'block'}
function toggleQuestions(){const e=document.getElementById('questionBox');if(e)e.style.display=e.style.display==='block'?'none':'block'}
function insuranceAnswer(sel,relatedIds){if(sel.value==='drop'){document.querySelectorAll('input[data-id]').forEach(x=>{if(relatedIds.includes(x.dataset.id))x.checked=false});recalc()}else if(sel.value==='keep'){document.querySelectorAll('input[data-id]').forEach(x=>{if(relatedIds.includes(x.dataset.id))x.checked=true});recalc()}}
document.getElementById('extraToggle').onclick=()=>{const b=document.getElementById('extraBody');b.classList.toggle('open');document.getElementById('extraToggle').textContent=b.classList.contains('open')?'Скрыть':'Разобраться'};

function render(){
  document.getElementById('results').style.display='block';
  document.getElementById('resultMeta').textContent=`Найдено ${result.candidates_count} потенциальных операций на ${rub(result.candidates_amount)} из ${result.transactions_scanned.toLocaleString('ru-RU')} банковских операций.`;
  document.getElementById('groups').innerHTML=result.groups.map(g=>`<span class="chip">${g.emoji} ${g.name}: ${g.count} · ${rub(g.amount)}</span>`).join('');
  document.getElementById('foundLabel').textContent=`${result.candidates_count} операций · ${rub(result.candidates_amount)}`;
  const byYear={};result.candidates.forEach(c=>(byYear[c.year]??=[]).push(c));let html='';
  Object.keys(byYear).sort().forEach(y=>{html+=`<tr class="yearhead"><td colspan="6">${y} год</td></tr>`;byYear[y].forEach(c=>{html+=`<tr><td><input type="checkbox" data-id="${c.id}" ${c.selected?'checked':''}></td><td>${c.date}</td><td>${c.emoji} ${c.category_name}</td><td class="merchant" title="${escapeHtml(c.note)}">${escapeHtml(c.merchant)}</td><td class="amt">${rub(c.amount)}</td><td><span class="status ${c.confidence==='high'?'high':'verify'}">${c.status}</span></td></tr>`})});
  document.getElementById('tbody').innerHTML=html;document.querySelectorAll('input[data-id]').forEach(x=>x.onchange=recalc);renderActions();document.getElementById('footer').textContent=`Файл: ${result.filename} · распознаватель: ${result.parser}`;document.getElementById('results').scrollIntoView({behavior:'smooth',block:'start'})
}
async function recalc(){const r=await fetch('/api/recalculate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({candidates:result.candidates,selected_ids:ids()})});await r.json();renderActions()}
document.getElementById('selectAll').onclick=()=>{const boxes=[...document.querySelectorAll('input[data-id]')],all=boxes.every(x=>x.checked);boxes.forEach(x=>x.checked=!all);document.getElementById('selectAll').textContent=!all?'Снять все':'Отметить все';recalc()};
document.getElementById('packet').onclick=async()=>{const r=await fetch('/api/packet',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({filename:result.filename,candidates:result.candidates,selected_ids:ids()})});if(!r.ok)return alert('Не удалось сформировать пакет');const blob=await r.blob(),u=URL.createObjectURL(blob),a=document.createElement('a');a.href=u;a.download='tax_radar_packet.html';a.click();URL.revokeObjectURL(u)};
</script>
</body>
</html>'''

